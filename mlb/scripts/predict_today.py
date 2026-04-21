"""
Generate model predictions for today's upcoming MLB games.

Uses:
  1. All completed 2026 games to build rolling team features (no look-ahead)
  2. Today's probable pitchers from the schedule API
  3. The 2025-trained logistic regression model

Outputs JSON to stdout for mlb.html and writes a daily markdown report to disk.

Usage:
    python mlb/scripts/predict_today.py
    python mlb/scripts/predict_today.py --date 2026-04-15
"""

import sys
import json
import time
import pickle
import argparse
import csv
import os
from pathlib import Path
from collections import defaultdict, deque
from datetime import date, datetime, timezone, timedelta
from html import escape

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import requests
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Load .env from repo root
_env_path = Path(__file__).parent.parent.parent / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            import os
            os.environ.setdefault(k.strip(), v.strip())

sys.path.append(str(Path(__file__).parent.parent.parent))
from mlb.scripts.model import FEATURES
from mlb.scripts.feature_utils import (
    BALLPARK_FACTORS,
    FILL_ERA,
    FILL_WHIP,
    FILL_K9,
    IP_BLEND_THRESHOLD,
    aggregate_bullpen_from_pitchers,
    bullpen_features,
    neutral_bullpen_features,
    pitcher_features,
    pitcher_row_from_stat,
    safe_float,
    parse_ip,
)
from mlb.scripts.odds_utils import (
    best_moneyline,
    best_spread,
    collect_spread_options as collect_filtered_spread_options,
    no_vig_probs,
    standard_run_line_points,
)

MODEL_DIR = Path(__file__).parent.parent / "models"
PREDICTIONS_DIR = Path(__file__).parent.parent / "predictions"
MLB = "https://statsapi.mlb.com/api/v1"
SEASON = 2026
BANKROLL_EUR = 500.0
ACCA_BANKROLL_EUR = 100.0

# Spread model gate. Selection still requires the saved spread model to pass its
# validation checks; setting this True lets the pipeline prove whether RL is
# blocked by validation/filtering instead of being globally disabled.
USE_SPREAD_MODEL = True
SPREAD_DEBUG = os.getenv("SPREAD_DEBUG", "").lower() in {"1", "true", "yes", "on"}

ODDS_TEAM_MAP = {
    "Arizona Diamondbacks": "AZ",
    "Baltimore Orioles": "BAL",
    "Boston Red Sox": "BOS",
    "Chicago Cubs": "CHC",
    "Chicago White Sox": "CWS",
    "Cincinnati Reds": "CIN",
    "Cleveland Guardians": "CLE",
    "Colorado Rockies": "COL",
    "Detroit Tigers": "DET",
    "Houston Astros": "HOU",
    "Kansas City Royals": "KC",
    "Los Angeles Angels": "LAA",
    "Los Angeles Dodgers": "LAD",
    "Miami Marlins": "MIA",
    "Milwaukee Brewers": "MIL",
    "Minnesota Twins": "MIN",
    "New York Mets": "NYM",
    "New York Yankees": "NYY",
    "Oakland Athletics": "ATH",
    "Athletics": "ATH",
    "Philadelphia Phillies": "PHI",
    "Pittsburgh Pirates": "PIT",
    "San Diego Padres": "SD",
    "San Francisco Giants": "SF",
    "Seattle Mariners": "SEA",
    "St. Louis Cardinals": "STL",
    "Tampa Bay Rays": "TB",
    "Texas Rangers": "TEX",
    "Toronto Blue Jays": "TOR",
    "Washington Nationals": "WSH",
}

PREFERRED_BOOKMAKERS = [
    "paddypower",
    "skybet",
    "boylesports",
]

UK_BOOKMAKERS = {
    "sport888",
    "betfair_ex_uk",
    "betfair_sb_uk",
    "betvictor",
    "betway",
    "boylesports",
    "casumo",
    "coral",
    "grosvenor",
    "ladbrokes_uk",
    "leovegas",
    "livescorebet",
    "matchbook",
    "paddypower",
    "skybet",
    "smarkets",
    "unibet_uk",
    "virginbet",
    "williamhill",
}


def get(url, params=None):
    r = requests.get(url, params=params, timeout=20)
    r.raise_for_status()
    return r.json()


def fetch_completed(today: str) -> list[dict]:
    data = get(
        f"{MLB}/schedule",
        params={
            "sportId": 1,
            "startDate": "2026-03-26",
            "endDate": today,
            "gameType": "R",
            "hydrate": "probablePitcher,team,linescore",
        },
    )
    games = []
    for day in data.get("dates", []):
        for g in day.get("games", []):
            if g.get("status", {}).get("statusCode") != "F":
                continue
            away = g["teams"]["away"]
            home = g["teams"]["home"]
            a_score = away.get("score")
            h_score = home.get("score")
            if a_score is None or h_score is None:
                ls = g.get("linescore", {}).get("teams", {})
                a_score = ls.get("away", {}).get("runs")
                h_score = ls.get("home", {}).get("runs")
            if a_score is None or h_score is None:
                continue
            games.append(
                {
                    "game_pk": g["gamePk"],
                    "game_date": day["date"],
                    "home_team_id": home["team"]["id"],
                    "away_team_id": away["team"]["id"],
                    "home_team": home["team"].get("abbreviation"),
                    "away_team": away["team"].get("abbreviation"),
                    "home_score": int(h_score),
                    "away_score": int(a_score),
                    "home_win": int(h_score) > int(a_score),
                }
            )
    return games


def fetch_upcoming(target_date: str) -> list[dict]:
    data = get(
        f"{MLB}/schedule",
        params={
            "sportId": 1,
            "startDate": target_date,
            "endDate": target_date,
            "gameType": "R",
            "hydrate": "probablePitcher,team,linescore",
        },
    )
    games = []
    for day in data.get("dates", []):
        for g in day.get("games", []):
            status_obj = g.get("status", {})
            status_code = status_obj.get("statusCode", "")
            abstract_state = status_obj.get("abstractGameState", "Preview")
            if status_code == "F" or abstract_state == "Final":
                continue
            away = g["teams"]["away"]
            home = g["teams"]["home"]
            away_sp = away.get("probablePitcher", {})
            home_sp = home.get("probablePitcher", {})
            games.append(
                {
                    "game_pk": g["gamePk"],
                    "game_date": target_date,
                    "series_game_number": g.get("seriesGameNumber"),
                    "games_in_series": g.get("gamesInSeries"),
                    "home_team_id": home["team"]["id"],
                    "home_team": home["team"]["abbreviation"],
                    "home_name": home["team"]["name"],
                    "away_team_id": away["team"]["id"],
                    "away_team": away["team"]["abbreviation"],
                    "away_name": away["team"]["name"],
                    "home_sp_id": home_sp.get("id"),
                    "home_sp_name": home_sp.get("fullName", "TBD"),
                    "away_sp_id": away_sp.get("id"),
                    "away_sp_name": away_sp.get("fullName", "TBD"),
                    "status": status_obj.get("detailedState", "Scheduled"),
                    # isLive: game has started — live odds from The Odds API must not be
                    # used for edge calculations (they are state-dependent, not market signals)
                    "isLive": abstract_state == "Live",
                }
            )
    return games


def fetch_pitcher_stats() -> dict:
    rows = {}
    offset, limit = 0, 500
    while True:
        data = get(
            f"{MLB}/stats",
            params={
                "stats": "season",
                "group": "pitching",
                "season": SEASON,
                "playerPool": "all",
                "limit": limit,
                "offset": offset,
                "hydrate": "person",
            },
        )
        splits = data.get("stats", [{}])[0].get("splits", [])
        if not splits:
            break
        for s in splits:
            p = s.get("player") or s.get("person") or {}
            st = s.get("stat", {})
            team = s.get("team") or {}
            pid = p.get("id")
            if not pid:
                continue
            row = pitcher_row_from_stat(st, p, team)
            row["name"] = row.get("pitcher_name", "?")
            rows[pid] = row
        if len(splits) < limit:
            break
        offset += limit
        time.sleep(0.2)
    return rows


def fetch_bullpen_stats(pitchers: dict) -> dict:
    """Aggregate current season bullpen/team pitching quality by team."""
    rows = []
    for row in pitchers.values():
        if row.get("team"):
            rows.append(row)
    if rows:
        df = aggregate_bullpen_from_pitchers(pd.DataFrame(rows))
        if not df.empty:
            return {r["team"]: r for r in df.to_dict("records")}

    # MLB player stat rows often omit team abbreviation. Fall back to team pitching
    # quality as a conservative bullpen proxy, then layer recent bullpen usage on top.
    try:
        teams = get(f"{MLB}/teams", params={"sportId": 1, "season": SEASON}).get("teams", [])
        id_to_abbr = {t["id"]: t.get("abbreviation") for t in teams}
        data = get(
            f"{MLB}/teams/stats",
            params={"stats": "season", "group": "pitching", "season": SEASON, "sportIds": 1},
        )
    except Exception as exc:
        print(f"// Bullpen quality fetch failed: {exc}", file=sys.stderr)
        return {}

    out = {}
    for split in data.get("stats", [{}])[0].get("splits", []):
        team_id = split.get("team", {}).get("id")
        abbr = id_to_abbr.get(team_id)
        if not abbr:
            continue
        st = split.get("stat", {})
        out[abbr] = {
            "team": abbr,
            "bp_era": safe_float(st.get("era"), 4.30),
            "bp_whip": safe_float(st.get("whip"), 1.32),
            "bp_k_bb": safe_float(st.get("strikeoutWalkRatio"), 2.45),
            "bp_ip_last_3d": 0.0,
            "bp_ip_yesterday": 0.0,
            "bp_relievers_last_3d": 0.0,
            "bp_relievers_yesterday": 0.0,
            "bp_top_used_yesterday": 0.0,
        }
    return out


def _compute_top_relievers_by_team(pitchers: dict) -> dict:
    """
    Returns {team_abbr: {"top2": set of pids, "top3": set of pids}}
    Uses pregame season-to-date stats only. Ranks by:
      saves*2 + holds*1 + k_bb_pct*20  (leverage score)
    """
    team_relievers: dict[str, list] = defaultdict(list)
    for pid, p in pitchers.items():
        if int(safe_float(p.get("games_started"), 0) or 0) > 0:
            continue
        team = p.get("team")
        if not team:
            continue
        saves = safe_float(p.get("saves"), 0) or 0.0
        holds = safe_float(p.get("holds"), 0) or 0.0
        k = safe_float(p.get("k"), 0) or 0.0
        walks = safe_float(p.get("walks"), 0) or 0.0
        bf = max(safe_float(p.get("batters_faced"), 1) or 1.0, 1.0)
        k_bb_pct = (k - walks) / bf
        score = saves * 2.0 + holds * 1.0 + max(0.0, k_bb_pct) * 20.0
        team_relievers[team].append((pid, score))
    result = {}
    for team, ranked in team_relievers.items():
        sorted_pids = [pid for pid, _ in sorted(ranked, key=lambda x: x[1], reverse=True)]
        result[team] = {"top2": set(sorted_pids[:2]), "top3": set(sorted_pids[:3])}
    return result


def _fetch_live_boxscore(game_pk: int) -> dict:
    try:
        return get(f"{MLB}/game/{game_pk}/feed/live")
    except Exception:
        return {}


def fetch_recent_bullpen_usage(completed: list[dict], pitchers: dict, target_date: str) -> dict:
    """
    Add bullpen fatigue from the latest completed games.

    Uses recent boxscores for innings and reliever counts. Computes leverage-aware
    features (top 2/top 3 by saves+holds+K-BB%) alongside legacy features.
    If a boxscore is missing, neutral fatigue values are preserved.
    """
    target = datetime.strptime(target_date, "%Y-%m-%d").date()
    recent = []
    for g in completed:
        g_date = datetime.strptime(g["game_date"], "%Y-%m-%d").date()
        days_back = (target - g_date).days
        if 1 <= days_back <= 3:
            recent.append((days_back, g))

    usage: dict = defaultdict(lambda: {
        "bp_ip_last_3d": 0.0,
        "bp_ip_yesterday": 0.0,
        "bp_relievers_last_3d": 0.0,
        "bp_relievers_yesterday": 0.0,
        "bp_top_used_yesterday": 0.0,
        "bp_top2_used_yesterday": 0.0,
        "bp_top2_backtoback": 0.0,
        "bp_top3_outs_last_3d": 0.0,
    })

    # Per-team leverage ranking from current season-to-date stats
    top_by_team = _compute_top_relievers_by_team(pitchers)

    # Legacy global threshold: any reliever with saves+holds >= 3 (no starts)
    global_top_ids = {
        pid for pid, p in pitchers.items()
        if (safe_float(p.get("saves"), 0) or 0) + (safe_float(p.get("holds"), 0) or 0) >= 3
        and (safe_float(p.get("games_started"), 0) or 0) <= 0
    }

    # Track per-team per-day reliever PIDs and top-3 outs for back-to-back + outs features
    team_day_pids: dict[str, dict[int, set]] = defaultdict(lambda: defaultdict(set))
    team_day_top3_outs: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))

    for days_back, g in recent:
        feed = _fetch_live_boxscore(g["game_pk"])
        teams_box = feed.get("liveData", {}).get("boxscore", {}).get("teams", {})
        for side in ("home", "away"):
            team_abbr = g.get(f"{side}_team")
            team_box = teams_box.get(side, {})
            pitcher_ids = team_box.get("pitchers", [])
            players = team_box.get("players", {})
            if not team_abbr or len(pitcher_ids) <= 1:
                continue
            # First pitcher in MLB boxscore is the starter; everyone after is bullpen
            reliever_ids = pitcher_ids[1:]
            top3_ids = top_by_team.get(team_abbr, {}).get("top3", set())
            reliever_ip = 0.0
            top3_outs = 0.0
            for pid in reliever_ids:
                pdata = players.get(f"ID{pid}", {})
                pitching = pdata.get("stats", {}).get("pitching", {})
                ip = parse_ip(pitching.get("inningsPitched", 0))
                reliever_ip += ip
                if pid in top3_ids:
                    top3_outs += ip * 3.0
                team_day_pids[team_abbr][days_back].add(pid)
            team_day_top3_outs[team_abbr][days_back] += top3_outs
            usage[team_abbr]["bp_ip_last_3d"] += reliever_ip
            usage[team_abbr]["bp_relievers_last_3d"] += len(reliever_ids)
            if days_back == 1:
                usage[team_abbr]["bp_ip_yesterday"] += reliever_ip
                usage[team_abbr]["bp_relievers_yesterday"] += len(reliever_ids)
                if any(pid in global_top_ids for pid in reliever_ids):
                    usage[team_abbr]["bp_top_used_yesterday"] = 1.0

    # Compute leverage features now that all days are collected
    all_teams = set(usage.keys()) | set(team_day_pids.keys())
    for team_abbr in all_teams:
        top2_ids = top_by_team.get(team_abbr, {}).get("top2", set())
        yday_pids = team_day_pids[team_abbr].get(1, set())
        day2_pids = team_day_pids[team_abbr].get(2, set())
        usage[team_abbr]["bp_top2_used_yesterday"] = 1.0 if top2_ids & yday_pids else 0.0
        # Requires the same reliever in both days (3-way intersection)
        usage[team_abbr]["bp_top2_backtoback"] = 1.0 if top2_ids & yday_pids & day2_pids else 0.0
        usage[team_abbr]["bp_top3_outs_last_3d"] = round(
            sum(team_day_top3_outs[team_abbr].values()), 1
        )

    return {team: dict(vals) for team, vals in usage.items()}


def build_team_state(completed: list[dict]) -> dict:
    team_history = defaultdict(lambda: deque(maxlen=20))
    for g in sorted(completed, key=lambda x: x["game_date"]):
        ht, at = g["home_team_id"], g["away_team_id"]
        hrd = g["home_score"] - g["away_score"]
        ard = -hrd
        team_history[ht].append((hrd, g["home_score"], g["away_score"], int(g["home_win"])))
        team_history[at].append((ard, g["away_score"], g["home_score"], int(not g["home_win"])))
    return {tid: list(hist) for tid, hist in team_history.items()}


def rolling(hist: list, n: int) -> dict:
    recent = hist[-n:]
    if len(recent) < 2:
        return {
            f"L{n}_WIN_PCT": None,
            f"L{n}_RD": None,
            f"L{n}_RUNS_FOR": None,
            f"L{n}_RUNS_AGN": None,
        }
    return {
        f"L{n}_WIN_PCT": round(float(np.mean([r[3] for r in recent])), 4),
        f"L{n}_RD": round(float(np.mean([r[0] for r in recent])), 4),
        f"L{n}_RUNS_FOR": round(float(np.mean([r[1] for r in recent])), 4),
        f"L{n}_RUNS_AGN": round(float(np.mean([r[2] for r in recent])), 4),
    }


def pitcher_stat(pitchers: dict, pid, col: str, fallback: float) -> float:
    if not pid or pid not in pitchers:
        return fallback
    v = pitchers[pid].get(col)
    return fallback if v is None else v


def blended_stat(pitchers: dict, pid, col: str, fallback: float) -> float:
    """Return a sample-size-adjusted stat for model features.

    Pitchers below IP_BLEND_THRESHOLD are blended toward the league-average
    fill value so that a 1-start ERA of 0.43 doesn't generate a false 20%+ edge.

        weight  = min(1.0, IP / IP_BLEND_THRESHOLD)
        blended = weight * raw + (1 - weight) * fallback

    Examples at FILL_ERA=4.50, threshold=30:
        Wacha  0.43 ERA,  10 IP  →  blended 3.14
        Sugano 2.16 ERA,   8 IP  →  blended 3.66
        Webb   5.25 ERA,  20 IP  →  blended 5.00
        Senga  7.07 ERA,  35 IP  →  fully trusted 7.07
    """
    if not pid or pid not in pitchers:
        return fallback
    p = pitchers[pid]
    v = p.get(col)
    if v is None:
        return fallback
    ip = p.get("ip", 0.0)
    weight = min(1.0, ip / IP_BLEND_THRESHOLD)
    return round(weight * v + (1.0 - weight) * fallback, 3)


def predict(model, scaler, feat_vec: list) -> float:
    X = np.array([feat_vec])
    return float(model.predict_proba(scaler.transform(X))[0][1])


def _feature_default(name: str) -> float:
    if name.endswith("_ERA"):
        return FILL_ERA
    if name.endswith("_WHIP"):
        return FILL_WHIP
    if name.endswith("_K9"):
        return FILL_K9
    if "WIN_PCT" in name:
        return 0.5
    if "RUNS_FOR" in name or "RUNS_AGN" in name:
        return 4.4
    if "BALLPARK" in name:
        return 1.0
    if "K_BB_PCT" in name:
        return 0.145
    if "BP_K_BB" in name:
        return 2.45
    return 0.0


def build_features(
    game: dict,
    team_state: dict,
    pitchers: dict,
    feature_list=None,
    bullpens: dict | None = None,
    bullpen_usage: dict | None = None,
) -> tuple[list, dict, bool]:
    if feature_list is None:
        feature_list = FEATURES

    ht = game["home_team_id"]
    at = game["away_team_id"]
    hh = team_state.get(ht, [])
    ah = team_state.get(at, [])

    hl5  = rolling(hh, 5)
    hl10 = rolling(hh, 10)
    hl20 = rolling(hh, 20)
    al5  = rolling(ah, 5)
    al10 = rolling(ah, 10)
    al20 = rolling(ah, 20)

    home_sp = pitcher_features(pitchers, game.get("home_sp_id"))
    away_sp = pitcher_features(pitchers, game.get("away_sp_id"))
    home_bp = bullpen_features(bullpens, game.get("home_team", ""))
    away_bp = bullpen_features(bullpens, game.get("away_team", ""))
    home_usage = (bullpen_usage or {}).get(game.get("home_team", ""), {})
    away_usage = (bullpen_usage or {}).get(game.get("away_team", ""), {})
    for key, value in home_usage.items():
        home_bp[key.upper()] = value
    for key, value in away_usage.items():
        away_bp[key.upper()] = value

    fd = {
        "HOME_L10_WIN_PCT":  hl10["L10_WIN_PCT"],
        "AWAY_L10_WIN_PCT":  al10["L10_WIN_PCT"],
        "HOME_L5_WIN_PCT":   hl5["L5_WIN_PCT"],
        "AWAY_L5_WIN_PCT":   al5["L5_WIN_PCT"],
        "HOME_L20_WIN_PCT":  hl20["L20_WIN_PCT"],
        "AWAY_L20_WIN_PCT":  al20["L20_WIN_PCT"],
        "HOME_L10_RD":       hl10["L10_RD"],
        "AWAY_L10_RD":       al10["L10_RD"],
        "HOME_L5_RD":        hl5["L5_RD"],
        "AWAY_L5_RD":        al5["L5_RD"],
        "HOME_L20_RD":       hl20["L20_RD"],
        "AWAY_L20_RD":       al20["L20_RD"],
        "HOME_L10_RUNS_FOR": hl10["L10_RUNS_FOR"],
        "AWAY_L10_RUNS_FOR": al10["L10_RUNS_FOR"],
        "HOME_L10_RUNS_AGN": hl10["L10_RUNS_AGN"],
        "AWAY_L10_RUNS_AGN": al10["L10_RUNS_AGN"],
        "HOME_L20_RUNS_FOR": hl20["L20_RUNS_FOR"],
        "AWAY_L20_RUNS_FOR": al20["L20_RUNS_FOR"],
        "HOME_L20_RUNS_AGN": hl20["L20_RUNS_AGN"],
        "AWAY_L20_RUNS_AGN": al20["L20_RUNS_AGN"],
        "BALLPARK_FACTOR":   BALLPARK_FACTORS.get(game.get("home_team", ""), 1.0),
    }
    for key, value in home_sp.items():
        fd[f"HOME_{key}"] = value
    for key, value in away_sp.items():
        fd[f"AWAY_{key}"] = value
    for key, value in home_bp.items():
        fd[f"HOME_{key}"] = value
    for key, value in away_bp.items():
        fd[f"AWAY_{key}"] = value

    fd["WIN_PCT_DIFF"] = (
        hl10["L10_WIN_PCT"] - al10["L10_WIN_PCT"]
        if hl10["L10_WIN_PCT"] is not None and al10["L10_WIN_PCT"] is not None
        else None
    )
    fd["RD_DIFF"] = hl10["L10_RD"] - al10["L10_RD"] if hl10["L10_RD"] is not None else None
    fd["WIN_PCT_DIFF_L20"] = (
        hl20["L20_WIN_PCT"] - al20["L20_WIN_PCT"]
        if hl20["L20_WIN_PCT"] is not None and al20["L20_WIN_PCT"] is not None
        else None
    )
    fd["RD_DIFF_L20"] = hl20["L20_RD"] - al20["L20_RD"] if hl20["L20_RD"] is not None else None
    fd["ERA_DIFF"] = fd["HOME_SP_ERA"] - fd["AWAY_SP_ERA"]
    fd["WHIP_DIFF"] = fd["HOME_SP_WHIP"] - fd["AWAY_SP_WHIP"]
    fd["K9_DIFF"] = fd["HOME_SP_K9"] - fd["AWAY_SP_K9"]
    fd["FIP_DIFF"] = fd["HOME_SP_FIP"] - fd["AWAY_SP_FIP"]
    fd["BB9_DIFF"] = fd["HOME_SP_BB9"] - fd["AWAY_SP_BB9"]
    fd["K_BB_PCT_DIFF"] = fd["HOME_SP_K_BB_PCT"] - fd["AWAY_SP_K_BB_PCT"]
    fd["HR9_DIFF"] = fd["HOME_SP_HR9"] - fd["AWAY_SP_HR9"]
    fd["SP_IP_DIFF"] = fd["HOME_SP_IP"] - fd["AWAY_SP_IP"]
    fd["SP_HAND_DIFF"] = fd["HOME_SP_IS_LEFT"] - fd["AWAY_SP_IS_LEFT"]
    fd["BP_ERA_DIFF"] = fd["HOME_BP_ERA"] - fd["AWAY_BP_ERA"]
    fd["BP_WHIP_DIFF"] = fd["HOME_BP_WHIP"] - fd["AWAY_BP_WHIP"]
    fd["BP_K_BB_DIFF"] = fd["HOME_BP_K_BB"] - fd["AWAY_BP_K_BB"]
    fd["BP_IP_LAST_3D_DIFF"] = fd["HOME_BP_IP_LAST_3D"] - fd["AWAY_BP_IP_LAST_3D"]
    fd["BP_IP_YESTERDAY_DIFF"] = fd["HOME_BP_IP_YESTERDAY"] - fd["AWAY_BP_IP_YESTERDAY"]
    fd["BP_RELIEVERS_LAST_3D_DIFF"] = fd["HOME_BP_RELIEVERS_LAST_3D"] - fd["AWAY_BP_RELIEVERS_LAST_3D"]
    fd["BP_RELIEVERS_YESTERDAY_DIFF"] = fd["HOME_BP_RELIEVERS_YESTERDAY"] - fd["AWAY_BP_RELIEVERS_YESTERDAY"]
    fd["BP_TOP_USED_YESTERDAY_DIFF"] = fd["HOME_BP_TOP_USED_YESTERDAY"] - fd["AWAY_BP_TOP_USED_YESTERDAY"]
    fd["BP_TOP2_USED_YESTERDAY_DIFF"] = fd.get("HOME_BP_TOP2_USED_YESTERDAY", 0.0) - fd.get("AWAY_BP_TOP2_USED_YESTERDAY", 0.0)
    fd["BP_TOP2_BACKTOBACK_DIFF"] = fd.get("HOME_BP_TOP2_BACKTOBACK", 0.0) - fd.get("AWAY_BP_TOP2_BACKTOBACK", 0.0)
    fd["BP_TOP3_OUTS_LAST_3D_DIFF"] = fd.get("HOME_BP_TOP3_OUTS_LAST_3D", 0.0) - fd.get("AWAY_BP_TOP3_OUTS_LAST_3D", 0.0)

    has_rolling = all(fd.get(f) is not None for f in feature_list)
    feat_vec = [fd.get(f) if fd.get(f) is not None else _feature_default(f) for f in feature_list]
    return feat_vec, fd, has_rolling


def stake_tier(edge: float, bankroll: float = BANKROLL_EUR) -> dict:
    if edge < 0.01:
        return {"pct": "0%", "pctValue": 0, "eur": 0.0, "label": "pass", "reportLabel": "PASS"}
    elif edge < 0.03:
        pct = 0.005
        label = "micro"
    elif edge < 0.06:
        pct = 0.01
        label = "low"
    elif edge < 0.10:
        pct = 0.02
        label = "low-mid"
    elif edge < 0.15:
        pct = 0.03
        label = "mid"
    elif edge < 0.20:
        pct = 0.04
        label = "mid-high"
    else:
        pct = 0.05
        label = "high"

    eur = round(bankroll * pct)          # rounded to nearest EUR for clean display
    pct_display = f"{pct * 100:.1f}".rstrip("0").rstrip(".")
    return {
        "pct": f"{pct_display}%",
        "pctValue": round(pct * 100, 1),
        "eur": float(eur),
        "label": label,
        "reportLabel": f"{pct_display}% (EUR {eur:.2f})",
    }


def ordinal(n: int) -> str:
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def format_prediction_date(target_date: str) -> tuple[str, str, str]:
    dt = datetime.strptime(target_date, "%Y-%m-%d")
    month_folder = f"{dt.strftime('%B')} Predictions"
    day_folder = f"{dt.strftime('%B')} {ordinal(dt.day)}"
    file_stub = f"{dt.strftime('%B')} {ordinal(dt.day)} {dt.year} Predictions"
    return month_folder, day_folder, file_stub


def moneyline_label(row: dict) -> str:
    if row.get("pickSide") == "home":
        return f'{row["homeAbbr"]} Moneyline'
    if row.get("pickSide") == "away":
        return f'{row["awayAbbr"]} Moneyline'
    return "SKIP"


def decision_for_row(row: dict) -> tuple[str, str]:
    if row.get("pickSide") == "none":
        return "SKIP", "SKIP - No clear edge"
    # Skip when either starting pitcher is unannounced — the pitching comparison
    # is incomplete and the model edge is unreliable.
    if row.get("homeSpName", "TBD") == "TBD" or row.get("awaySpName", "TBD") == "TBD":
        return "SKIP", "SKIP - SP not yet announced"
    has_odds = bool(row.get("hasOdds")) and row.get("pickOdds") is not None
    if not has_odds:
        return "SKIP", "SKIP - Missing odds data"
    if row.get("edge", 0.0) < 0.01 or row["stake"]["eur"] <= 0:
        return "SKIP", "SKIP - No clear edge"
    if row.get("useRl"):
        pick_abbr = row["homeAbbr"] if row.get("pickSide") == "home" else row["awayAbbr"]
        sp = row.get("spreadPoint")
        sp_label = f"{sp:+g}" if sp is not None else "-1.5"
        return "BET", f"{pick_abbr} {sp_label} (Run Line)"
    return "BET", moneyline_label(row)


def build_reasons(row: dict) -> list[str]:
    reasons = []
    edge = row.get("edge", 0.0)
    if edge > 0:
        reasons.append(f'Model edge of {edge * 100:.1f}% versus market implied probability.')

    pick_side = row.get("pickSide")
    home_l10wp = row.get("homeL10WP")
    away_l10wp = row.get("awayL10WP")
    home_l10rd = row.get("homeL10RD")
    away_l10rd = row.get("awayL10RD")
    home_era = row.get("homeSpEra")
    away_era = row.get("awaySpEra")
    home_whip = row.get("homeSpWhip")
    away_whip = row.get("awaySpWhip")

    if pick_side == "home":
        if home_l10wp is not None and away_l10wp is not None and home_l10wp > away_l10wp:
            reasons.append("Home team has the stronger recent L10 win rate.")
        if home_l10rd is not None and away_l10rd is not None and home_l10rd > away_l10rd:
            reasons.append("Home team has the better recent run differential.")
        if home_era is not None and away_era is not None and home_era < away_era:
            reasons.append("Home starter has the better ERA matchup.")
        if home_whip is not None and away_whip is not None and home_whip < away_whip:
            reasons.append("Home starter has the better WHIP profile.")
    elif pick_side == "away":
        if away_l10wp is not None and home_l10wp is not None and away_l10wp > home_l10wp:
            reasons.append("Away team has the stronger recent L10 win rate.")
        if away_l10rd is not None and home_l10rd is not None and away_l10rd > home_l10rd:
            reasons.append("Away team has the better recent run differential.")
        if away_era is not None and home_era is not None and away_era < home_era:
            reasons.append("Away starter has the better ERA matchup.")
        if away_whip is not None and home_whip is not None and away_whip < home_whip:
            reasons.append("Away starter has the better WHIP profile.")

    if not reasons:
        reasons.append("No strong statistical edge beyond model price value.")
    return reasons[:3]


def build_risk(row: dict) -> str:
    home_prob = row.get("homeProb")
    away_prob = row.get("awayProb")
    edge = row.get("edge", 0.0)
    pick_side = row.get("pickSide")
    home_era = row.get("homeSpEra")
    away_era = row.get("awaySpEra")
    home_ip = row.get("homeSpIp", 0.0) or 0.0
    away_ip = row.get("awaySpIp", 0.0) or 0.0

    # Warn when model is working with limited pitcher data
    if home_ip < IP_BLEND_THRESHOLD and home_ip > 0:
        return f"Home SP has only {home_ip:.1f} IP — ERA blended toward league average ({FILL_ERA})."
    if away_ip < IP_BLEND_THRESHOLD and away_ip > 0:
        return f"Away SP has only {away_ip:.1f} IP — ERA blended toward league average ({FILL_ERA})."

    if home_prob is not None and away_prob is not None and abs(home_prob - away_prob) < 0.06:
        return "Model sees this as a fairly close game."
    if pick_side == "home" and home_era is not None and away_era is not None and home_era > away_era:
        return "Picked side does not have the better ERA matchup."
    if pick_side == "away" and away_era is not None and home_era is not None and away_era > home_era:
        return "Picked side does not have the better ERA matchup."
    if edge < 0.05:
        return "Edge is modest, so variance risk is higher."
    return "Baseball variance is high even with a real edge."


def confidence_for_row(row: dict) -> int:
    return min(10, max(1, int(round(row.get("edge", 0.0) * 100 / 2))))


def build_summary_rows(predictions: list[dict]) -> list[dict]:
    rows = []
    for row in predictions:
        decision, pick_text = decision_for_row(row)
        use_rl = row.get("useRl", False)
        rl_pick_odds = row.get("rlPickOdds")
        ml_odds = row.get("pickOdds")
        # Use RL odds for P&L when RL is the primary bet
        odds = rl_pick_odds if use_rl and rl_pick_odds else ml_odds
        odds_text = f"{odds:.2f}" if isinstance(odds, (int, float)) else "N/A"
        stake_pct = row["stake"].get("pctValue", 0)
        stake_eur = float(row["stake"].get("eur", 0.0))
        if decision == "BET" and isinstance(odds, (int, float)):
            gross_return = round(stake_eur * odds, 2)
            profit = round(gross_return - stake_eur, 2)
        else:
            gross_return = 0.00
            profit = 0.00
        rows.append({
            "game": f'{row["awayAbbr"]} @ {row["homeAbbr"]}',
            "pick": pick_text,
            "odds_text": odds_text,
            "stake_pct": stake_pct,
            "stake_eur": stake_eur,
            "gross_return": gross_return,
            "profit": profit,
            "decision": decision,
        })
    return rows


def write_csv_report(predictions: list[dict], target_date: str) -> Path:
    month_folder, day_folder, file_stub = format_prediction_date(target_date)
    out_dir = PREDICTIONS_DIR / month_folder / day_folder
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{file_stub}.csv"

    with out_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Game",
            "Series",
            "Pick",
            "Odds",
            "Stake %",
            "Stake EUR",
            "Return EUR",
            "Profit EUR",
            "Decision",
            "Confidence",
            "Top Reasons",
            "Risk",
            "Book",
        ])
        for row, summary in zip(predictions, build_summary_rows(predictions)):
            series = ""
            if row.get("seriesGameNumber") and row.get("gamesInSeries"):
                series = f'Game {row["seriesGameNumber"]} of {row["gamesInSeries"]}'
            book = row.get("homeBook") if row.get("pickSide") == "home" else row.get("awayBook")
            writer.writerow([
                f'{row["awayTeam"]} vs {row["homeTeam"]}',
                series,
                summary["pick"],
                summary["odds_text"],
                f'{summary["stake_pct"]}%',
                f'EUR {summary["stake_eur"]:.2f}',
                f'EUR {summary["gross_return"]:.2f}',
                f'EUR {summary["profit"]:.2f}',
                summary["decision"],
                confidence_for_row(row),
                " | ".join(build_reasons(row)),
                build_risk(row),
                book or "",
            ])
    return out_file


def format_acca_section(accumulators: list[dict]) -> list[str]:
    """Generate the accumulator / bet-builder section for the markdown report."""
    if not accumulators:
        return []
    lines = [
        "",
        "---",
        "",
        f"## Acca / Bet Builder  *(Acca Bankroll: EUR {ACCA_BANKROLL_EUR:.0f} — separate from singles)*",
        "",
        "> Staked from the dedicated acca bankroll only. "
        "Singles picks, stakes, and P&L are unaffected. "
        "All legs are backed by real odds from The Odds API. "
        "Track acca results separately.",
        "",
    ]
    for acca in accumulators:
        lines.append(f"### {acca['type']}  —  Combined odds: {acca['combined_odds']:.2f}")
        for i, leg in enumerate(acca["legs"], 1):
            lines.append(
                f"{i}. {leg['game']}  →  **{leg['label']}** @ {leg['odds']:.2f}"
                f"  *(model edge: {leg['edge']*100:.1f}%)*"
            )
        lines.append(
            f"**Acca stake:** EUR {acca['stake']:.2f}  |  "
            f"**Potential return:** EUR {acca['potential_return']:.2f}"
        )
        lines.append("")
    return lines


def write_markdown_report(predictions: list[dict], target_date: str, bankroll: float = BANKROLL_EUR, odds_fetched_at: str = "N/A", accumulators: list[dict] | None = None) -> Path:
    month_folder, day_folder, file_stub = format_prediction_date(target_date)
    out_dir = PREDICTIONS_DIR / month_folder / day_folder
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{file_stub}.md"

    dt = datetime.strptime(target_date, "%Y-%m-%d")
    pretty_date = f'{dt.strftime("%B")} {ordinal(dt.day)} {dt.year} Predictions'
    divider = "━" * 23

    lines = [
        f"# {pretty_date}",
        "",
        f"**Bankroll:** EUR {bankroll:.2f}",
        f"**Generated for:** {target_date}",
        f"**Odds fetched:** {odds_fetched_at}",
        "",
    ]

    summary_rows = build_summary_rows(predictions)
    for row, summary in zip(predictions, summary_rows):
        decision, pick_text = decision_for_row(row)
        odds = row.get("pickOdds")
        odds_text = f"{odds:.2f}" if isinstance(odds, (int, float)) else "N/A"
        confidence = confidence_for_row(row)
        reasons = build_reasons(row)
        risk = build_risk(row)
        stake_pct = row["stake"].get("pctValue", 0)
        stake_eur = float(row["stake"].get("eur", 0.0))
        edge = row.get("edge", 0.0)

        series = ""
        if row.get("seriesGameNumber") and row.get("gamesInSeries"):
            series = f' (Game {row["seriesGameNumber"]} of {row["gamesInSeries"]})'

        # Pitcher display
        home_sp = row.get("homeSpName", "TBD")
        away_sp = row.get("awaySpName", "TBD")
        home_era  = row.get("homeSpEra")
        home_whip = row.get("homeSpWhip")
        away_era  = row.get("awaySpEra")
        away_whip = row.get("awaySpWhip")
        home_ip = row.get("homeSpIp", 0.0) or 0.0
        away_ip = row.get("awaySpIp", 0.0) or 0.0

        home_sp_str = home_sp
        if home_era is not None:
            home_sp_str += f" (ERA {home_era:.2f}"
            if home_whip is not None:
                home_sp_str += f" / WHIP {home_whip:.2f}"
            home_sp_str += ")"
        if 0 < home_ip < IP_BLEND_THRESHOLD:
            home_sp_str += f" ⚠️ {home_ip:.1f} IP"

        away_sp_str = away_sp
        if away_era is not None:
            away_sp_str += f" (ERA {away_era:.2f}"
            if away_whip is not None:
                away_sp_str += f" / WHIP {away_whip:.2f}"
            away_sp_str += ")"
        if 0 < away_ip < IP_BLEND_THRESHOLD:
            away_sp_str += f" ⚠️ {away_ip:.1f} IP"

        # Run line
        home_rl = row.get("homeRl")
        away_rl = row.get("awayRl")
        use_rl = row.get("useRl", False)
        rl_pick_odds = row.get("rlPickOdds")
        rl_line = ""
        if home_rl is not None or away_rl is not None:
            rl_parts = []
            if home_rl is not None:
                rl_parts.append(f'{row["homeAbbr"]} {row.get("homeRlPoint", -1.5):+g} @ {home_rl:.2f}')
            if away_rl is not None:
                rl_parts.append(f'{row["awayAbbr"]} {row.get("awayRlPoint", 1.5):+g} @ {away_rl:.2f}')
            rl_line = "📐 RUN LINE: " + "  |  ".join(rl_parts)

        # When RL is the primary bet, use RL odds for display
        display_odds = rl_pick_odds if use_rl and rl_pick_odds else odds
        display_odds_text = f"{display_odds:.2f}" if isinstance(display_odds, (int, float)) else "N/A"

        lines.append(divider)
        lines.append(f'⚾ GAME: {row["awayTeam"]} vs {row["homeTeam"]}{series}')
        lines.append(f'🏟️  Home SP: {home_sp_str}')
        lines.append(f'✈️  Away SP: {away_sp_str}')
        lines.append("")
        lines.append(f"📊 PICK: {pick_text}")
        lines.append(f"💰 STAKE: {stake_pct}% (EUR {stake_eur:.2f})")
        if use_rl and rl_pick_odds:
            lines.append(f"📈 ODDS: {display_odds_text}  (ML was {odds_text} — RL selected for better value)")
        else:
            lines.append(f"📈 ODDS: {odds_text}")
        if rl_line:
            lines.append(rl_line)
        lines.append("")
        lines.append(f"🔥 CONFIDENCE: {confidence}")
        lines.append("")
        lines.append("🧠 EDGE:")
        for r in reasons:
            lines.append(f"- {r}")
        lines.append("")
        lines.append("📉 RISKS:")
        lines.append(f"- {risk}")
        if decision == "BET" and edge < 0.03:
            lines.append(f"⚠️  TIGHT EDGE: {edge * 100:.1f}% — thin margin, treat as Tier C volume play. Check odds before placing.")
        lines.append("")
        lines.append(f"Decision: {decision}")
        lines.append(divider)
        lines.append("")

    # Summary table
    lines.append("## Summary Table")
    lines.append("")
    lines.append("| Game | Pick | Odds | Stake % | Stake EUR | Return EUR | Profit EUR | Decision |")
    lines.append("|---|---|---:|---:|---:|---:|---:|---|")
    for summary in summary_rows:
        lines.append(
            f'| {summary["game"]} '
            f'| {summary["pick"]} '
            f'| {summary["odds_text"]} '
            f'| {summary["stake_pct"]}% '
            f'| EUR {summary["stake_eur"]:.2f} '
            f'| EUR {summary["gross_return"]:.2f} '
            f'| EUR {summary["profit"]:.2f} '
            f'| {summary["decision"]} |'
        )

    lines.extend(format_acca_section(accumulators or []))

    out_file.write_text("\n".join(lines), encoding="utf-8")
    return out_file


# ─── ACCUMULATOR STAKING ──────────────────────────────────────────────────────
# Each combo type has its own fixed stake. Larger combos risk less per ticket
# but pay more if they land.
ACCA_TYPE_STAKES = {
    "Double": 10.00,
    "Treble": 10.00,
    "Quad":    7.00,
    "5-Fold":  5.00,
    "6-Fold":  3.00,
}
ACCA_STAKE = 10.00  # legacy fallback


def _best_acca_leg(row: dict) -> dict | None:
    """
    Find the best-value market for one game as an accumulator leg.

    Strategy:
    - Start with ML as the baseline (model directly predicts this market).
    - Evaluate ±1.5 run-line options using probability-adjusted edges.
    - Never use -2.5 or worse: the model doesn't predict cover probability
      reliably at large margins, and the risk kills multi-leg accas.
    - Pick the market with the highest ADJUSTED edge, not the highest odds.
    - Return None if no market has a genuine positive edge.
    """
    pick_side = row.get("pickSide")
    if not pick_side or pick_side == "none":
        return None

    model_prob = row.get("modelProb", 0.0)
    if model_prob < 0.52:          # slightly more permissive for accas
        return None

    abbr     = row["homeAbbr"] if pick_side == "home" else row["awayAbbr"]
    game     = f'{row["awayAbbr"]} @ {row["homeAbbr"]}'
    ml_odds  = row.get("homeMl") if pick_side == "home" else row.get("awayMl")

    if not isinstance(ml_odds, (int, float)) or ml_odds <= 1.0:
        return None

    # Baseline: ML
    ml_edge  = round(model_prob - 1.0 / ml_odds, 4)
    best = {
        "label":    f"{abbr} ML",
        "odds":     round(float(ml_odds), 2),
        "line":     "ml",
        "adj_prob": model_prob,
        "edge":     ml_edge,
    }

    # Run-line cover probability is not derived from the ML model anymore.
    # Accumulators stay on ML until a separate RL model is trained and saved.
    if best["edge"] <= 0:
        return None
    return {
        "gamePk":   str(row.get("gamePk", "")),
        "game":     game,
        "label":    best["label"],
        "odds":     best["odds"],
        "edge":     best["edge"],
        "adjProb":  round(best["adj_prob"], 4),
        "line":     "ml",
        "pickSide": pick_side,
    }

def build_accumulators(predictions: list[dict]) -> list[dict]:
    """
    Build a tiered suite of value-based accumulators from today's predictions.

    Framework:
    - Every leg must have genuine positive EV individually (edge ≥ 3%).
    - Larger combos (4+ legs) require every selected leg to clear 5% edge —
      no weak legs added just to boost combined odds.
    - Combined EV is computed from the product of adjusted probabilities vs
      the combined market implied probability.

    Stakes:
      Double  (top 2 legs, ≥3% each) — €10
      Treble  (top 3 legs, ≥3% each) — €10
      Quad    (top 4 legs, ≥5% each) — €7
      5-Fold  (top 5 legs, ≥5% each) — €5
      6-Fold  (top 6 legs, ≥5% each) — €3

    All legs use real API odds. Staking from ACCA_BANKROLL_EUR only.
    """
    # Gather all qualified legs (positive adjusted edge per game)
    raw_legs = []
    for row in predictions:
        leg = _best_acca_leg(row)
        if leg is not None:
            raw_legs.append(leg)

    # Sort by edge descending
    raw_legs.sort(key=lambda x: x["edge"], reverse=True)

    # Tier 1: legs with ≥3% adjusted edge — used for Double + Treble
    legs_3pct = [l for l in raw_legs if l["edge"] >= 0.03]
    # Tier 2: legs with ≥5% adjusted edge — used for Quad, 5-Fold, 6-Fold
    legs_5pct = [l for l in raw_legs if l["edge"] >= 0.05]

    def _combined_odds(selection: list[dict]) -> float:
        result = 1.0
        for l in selection:
            result *= l["odds"]
        return round(result, 2)

    def _combined_ev(selection: list[dict]) -> float:
        """
        EV of the accumulator ticket itself.
        = (product of true probabilities) / (combined market implied prob) - 1
        A positive value means the ticket has positive expected value.
        """
        true_prob   = 1.0
        market_prob = 1.0
        for l in selection:
            true_prob   *= l.get("adjProb", l["edge"] + 1.0 / l["odds"])
            market_prob *= 1.0 / l["odds"]
        return round(true_prob / market_prob - 1, 4) if market_prob > 0 else 0.0

    accas = []

    # Double + Treble from best ≥3% legs
    for type_name, n_legs, stake in [("Double", 2, 10.00), ("Treble", 3, 10.00)]:
        if len(legs_3pct) < n_legs:
            continue
        selection = legs_3pct[:n_legs]
        combined  = _combined_odds(selection)
        accas.append({
            "type":             type_name,
            "legs":             selection,
            "combined_odds":    combined,
            "stake":            stake,
            "potential_return": round(stake * combined, 2),
            "combined_ev":      _combined_ev(selection),
        })

    # Quad / 5-Fold / 6-Fold from best ≥5% legs only
    for type_name, n_legs, stake in [("Quad", 4, 7.00), ("5-Fold", 5, 5.00), ("6-Fold", 6, 3.00)]:
        if len(legs_5pct) < n_legs:
            continue
        selection = legs_5pct[:n_legs]
        combined  = _combined_odds(selection)
        accas.append({
            "type":             type_name,
            "legs":             selection,
            "combined_odds":    combined,
            "stake":            stake,
            "potential_return": round(stake * combined, 2),
            "combined_ev":      _combined_ev(selection),
        })

    return accas


def write_json_report(predictions: list[dict], target_date: str, accumulators: list[dict] | None = None) -> Path:
    month_folder, day_folder, file_stub = format_prediction_date(target_date)
    out_dir = PREDICTIONS_DIR / month_folder / day_folder
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{file_stub}.json"

    payload = {
        "date": target_date,
        "bankroll": BANKROLL_EUR,
        "snapshotKind": os.getenv("PREDICTION_SNAPSHOT_KIND", "morning"),
        "snapshotNote": os.getenv("PREDICTION_SNAPSHOT_NOTE", ""),
        "regenerated": os.getenv("PREDICTION_REGENERATED", "").lower() in {"1", "true", "yes", "on"},
        "predictions": predictions,
        "accumulators": accumulators or [],
    }
    out_file.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return out_file


def write_excel_report(predictions: list[dict], target_date: str, bankroll: float = BANKROLL_EUR) -> Path:
    month_folder, day_folder, file_stub = format_prediction_date(target_date)
    out_dir = PREDICTIONS_DIR / month_folder / day_folder
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{file_stub}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"

    title_fill = PatternFill("solid", fgColor="183153")
    sub_fill = PatternFill("solid", fgColor="E7F0FF")
    header_fill = PatternFill("solid", fgColor="244063")
    bet_fill = PatternFill("solid", fgColor="EAF8F1")
    skip_fill = PatternFill("solid", fgColor="FFF1EC")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D7E2"),
        right=Side(style="thin", color="D0D7E2"),
        top=Side(style="thin", color="D0D7E2"),
        bottom=Side(style="thin", color="D0D7E2"),
    )

    pretty_title = f"{format_prediction_date(target_date)[2]} Tracker"
    ws["A1"] = pretty_title
    ws["A2"] = f"Bankroll: EUR {bankroll:.2f}"
    ws["A3"] = f"Generated for: {target_date}"
    ws["A4"] = "Select Win / Loss / Push in the Result column and totals update automatically."
    for cell in ("A1", "A2", "A3", "A4"):
        ws[cell].font = title_font if cell == "A1" else bold_font
        ws[cell].fill = title_fill if cell == "A1" else sub_fill
        ws[cell].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:H1")

    headers = ["Game", "Pick", "Odds", "Stake", "Decision", "Result", "Profit/Loss", "Game PK"]

    header_row = 6
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_rows = build_summary_rows(predictions)
    for idx, (row, summary) in enumerate(zip(predictions, summary_rows), start=header_row + 1):
        series = ""
        if row.get("seriesGameNumber") and row.get("gamesInSeries"):
            series = f' (Game {row["seriesGameNumber"]} of {row["gamesInSeries"]})'
        stake_text = f"EUR {summary['stake_eur']:.2f}" if summary["stake_eur"] else "EUR 0.00"
        row_number = idx
        # Use RL odds for the Excel odds column when RL is the primary bet
        use_rl = row.get("useRl", False)
        rl_pick_odds = row.get("rlPickOdds")
        excel_odds = rl_pick_odds if use_rl and rl_pick_odds else (
            None if summary["odds_text"] == "N/A" else float(summary["odds_text"])
        )
        values = [
            f'{row["awayTeam"]} vs {row["homeTeam"]}{series}',
            summary["pick"],
            excel_odds,
            stake_text,
            summary["decision"],
            "",
            f'=IF(E{row_number}<>"BET",0,IF(F{row_number}="Win",(C{row_number}-1)*VALUE(SUBSTITUTE(D{row_number},"EUR ","")),IF(F{row_number}="Loss",-VALUE(SUBSTITUTE(D{row_number},"EUR ","")),IF(F{row_number}="Push",0,""))))',
            str(row.get("gamePk", "")),
        ]

        fill = bet_fill if summary["decision"] == "BET" else skip_fill
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.cell(row=idx, column=3).number_format = "0.00"
        ws.cell(row=idx, column=7).number_format = '"EUR" 0.00'
        ws.row_dimensions[idx].height = 15

    data_start = header_row + 1
    data_end = ws.max_row

    result_validation = DataValidation(type="list", formula1='"Win,Loss,Push"', allow_blank=True)
    result_validation.prompt = "Select a result"
    result_validation.error = "Choose Win, Loss, or Push"
    ws.add_data_validation(result_validation)
    result_validation.add(f"F{data_start}:F{data_end}")

    totals_row = data_end + 2
    ws[f"A{totals_row}"] = "Total Bet Count:"
    ws[f"B{totals_row}"] = f'=COUNTIF($E${data_start}:$E${data_end},"BET")'
    ws[f"A{totals_row + 1}"] = "Total Staked:"
    ws[f"B{totals_row + 1}"] = f'=SUMPRODUCT(--($E${data_start}:$E${data_end}="BET"),VALUE(SUBSTITUTE($D${data_start}:$D${data_end},"EUR ","")))'
    ws[f"A{totals_row + 2}"] = "Total Profit:"
    ws[f"B{totals_row + 2}"] = f"=SUM(G{data_start}:G{data_end})"
    ws[f"A{totals_row + 3}"] = "Net Bankroll:"
    ws[f"B{totals_row + 3}"] = f"={bankroll}+B{totals_row + 2}"

    for row_idx in range(totals_row, totals_row + 4):
        ws[f"A{row_idx}"].font = bold_font
        ws[f"A{row_idx}"].fill = sub_fill
        ws[f"B{row_idx}"].font = bold_font
        ws[f"B{row_idx}"].fill = sub_fill
        ws[f"A{row_idx}"].border = thin_border
        ws[f"B{row_idx}"].border = thin_border
        ws[f"B{row_idx}"].number_format = "0" if row_idx == totals_row else '"EUR" 0.00'
        ws.row_dimensions[row_idx].height = 15

    for title_row in range(1, header_row + 1):
        ws.row_dimensions[title_row].height = 15

    widths = {
        "A": 42,
        "B": 24,
        "C": 10,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 14,
        "H": 12,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    profit_green = PatternFill("solid", fgColor="D8F0DD")
    profit_red = PatternFill("solid", fgColor="F8D7DA")
    ws.conditional_formatting.add(
        f"B{totals_row + 2}:B{totals_row + 3}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=profit_green),
    )
    ws.conditional_formatting.add(
        f"B{totals_row + 2}:B{totals_row + 3}",
        CellIsRule(operator="lessThan", formula=["0"], fill=profit_red),
    )

    ws.freeze_panes = "A7"
    ws.column_dimensions["H"].hidden = True
    ws.auto_filter.ref = f"A6:H{data_end}"

    wb.save(out_file)
    return out_file


def write_html_report(predictions: list[dict], target_date: str, bankroll: float = BANKROLL_EUR) -> Path:
    month_folder, day_folder, file_stub = format_prediction_date(target_date)
    out_dir = PREDICTIONS_DIR / month_folder / day_folder
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{file_stub}.html"

    dt = datetime.strptime(target_date, "%Y-%m-%d")
    pretty_date = f'{dt.strftime("%B")} {ordinal(dt.day)} {dt.year} Predictions'
    summary_rows = build_summary_rows(predictions)

    cards = []
    for row in predictions:
        decision, pick_text = decision_for_row(row)
        odds = row.get("pickOdds")
        odds_text = f"{odds:.2f}" if isinstance(odds, (int, float)) else "N/A"
        confidence = confidence_for_row(row)
        reasons = "".join(f"<li>{escape(reason)}</li>" for reason in build_reasons(row))
        risk = escape(build_risk(row))
        stake_pct = row["stake"].get("pctValue", 0)
        stake_eur = float(row["stake"].get("eur", 0.0))
        game_title = f'{row["awayTeam"]} vs {row["homeTeam"]}'
        if row.get("seriesGameNumber") and row.get("gamesInSeries"):
            game_title += f' <span class="series">Game {row["seriesGameNumber"]} of {row["gamesInSeries"]}</span>'
        book = row.get("homeBook") if row.get("pickSide") == "home" else row.get("awayBook")
        book_label = book or "N/A"
        decision_class = "bet" if decision == "BET" else "skip"

        cards.append(
            f"""
            <section class="card {decision_class}">
              <div class="card-top">
                <h2>{game_title}</h2>
                <span class="decision {decision_class}">{decision}</span>
              </div>
              <div class="meta-grid">
                <div><span class="label">Pick</span><strong>{escape(pick_text)}</strong></div>
                <div><span class="label">Stake</span><strong>{stake_pct}% (EUR {stake_eur:.2f})</strong></div>
                <div><span class="label">Odds</span><strong>{odds_text}</strong></div>
                <div><span class="label">Confidence</span><strong>{confidence}/10</strong></div>
                <div><span class="label">Book</span><strong>{escape(book_label)}</strong></div>
                <div><span class="label">Edge</span><strong>{row.get("edge", 0.0) * 100:.1f}%</strong></div>
              </div>
              <div class="columns">
                <div>
                  <h3>Edge</h3>
                  <ul>{reasons}</ul>
                </div>
                <div>
                  <h3>Risk</h3>
                  <p>{risk}</p>
                </div>
              </div>
            </section>
            """
        )

    table_rows = []
    for summary in summary_rows:
        table_rows.append(
            f"""
            <tr>
              <td>{escape(summary["game"])}</td>
              <td>{escape(summary["pick"])}</td>
              <td>{summary["odds_text"]}</td>
              <td>{summary["stake_pct"]}%</td>
              <td>EUR {summary["stake_eur"]:.2f}</td>
              <td>EUR {summary["gross_return"]:.2f}</td>
              <td>EUR {summary["profit"]:.2f}</td>
              <td><span class="decision {summary["decision"].lower()}">{summary["decision"]}</span></td>
            </tr>
            """
        )

    html = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(pretty_date)}</title>
  <style>
    :root {{
      --bg: #0b1220;
      --panel: #111a2b;
      --panel-alt: #17233a;
      --text: #e7edf7;
      --muted: #91a0b8;
      --line: #24324d;
      --bet: #19c37d;
      --skip: #ff8a65;
      --accent: #6cb8ff;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", system-ui, sans-serif;
      background: radial-gradient(circle at top, #14203a, var(--bg) 48%);
      color: var(--text);
      line-height: 1.45;
    }}
    .wrap {{
      max-width: 1100px;
      margin: 0 auto;
      padding: 32px 20px 56px;
    }}
    .hero {{
      background: linear-gradient(135deg, rgba(108,184,255,.12), rgba(25,195,125,.08));
      border: 1px solid var(--line);
      border-radius: 20px;
      padding: 24px;
      margin-bottom: 24px;
    }}
    .hero h1 {{ margin: 0 0 8px; font-size: 2rem; }}
    .hero p {{ margin: 4px 0; color: var(--muted); }}
    .card {{
      background: linear-gradient(180deg, var(--panel), var(--panel-alt));
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 20px;
      margin-bottom: 18px;
      box-shadow: 0 10px 30px rgba(0,0,0,.18);
    }}
    .card.bet {{ border-color: rgba(25,195,125,.38); }}
    .card.skip {{ border-color: rgba(255,138,101,.28); }}
    .card-top {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      margin-bottom: 16px;
    }}
    .card-top h2 {{ margin: 0; font-size: 1.18rem; }}
    .series {{
      display: inline-block;
      margin-left: 8px;
      font-size: .78rem;
      color: var(--accent);
      font-weight: 600;
    }}
    .decision {{
      display: inline-block;
      padding: 6px 10px;
      border-radius: 999px;
      font-weight: 700;
      font-size: .82rem;
    }}
    .decision.bet {{ background: rgba(25,195,125,.16); color: var(--bet); }}
    .decision.skip {{ background: rgba(255,138,101,.14); color: var(--skip); }}
    .meta-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
      gap: 12px;
      margin-bottom: 16px;
    }}
    .meta-grid div {{
      background: rgba(255,255,255,.025);
      border: 1px solid rgba(255,255,255,.04);
      border-radius: 12px;
      padding: 10px 12px;
    }}
    .label {{
      display: block;
      color: var(--muted);
      font-size: .76rem;
      margin-bottom: 4px;
      text-transform: uppercase;
      letter-spacing: .04em;
    }}
    .columns {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
    }}
    .columns h3 {{
      margin: 0 0 8px;
      font-size: .92rem;
      color: var(--accent);
      text-transform: uppercase;
      letter-spacing: .04em;
    }}
    .columns ul, .columns p {{ margin: 0; color: var(--text); }}
    .columns ul {{ padding-left: 18px; }}
    .summary {{
      margin-top: 28px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 18px;
      overflow: hidden;
    }}
    .summary h2 {{
      margin: 0;
      padding: 18px 20px;
      border-bottom: 1px solid var(--line);
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: .95rem;
    }}
    th, td {{
      padding: 12px 14px;
      border-bottom: 1px solid var(--line);
      text-align: left;
    }}
    th {{ color: var(--muted); font-size: .82rem; text-transform: uppercase; }}
    tr:last-child td {{ border-bottom: none; }}
    @media (max-width: 760px) {{
      .columns {{ grid-template-columns: 1fr; }}
      .card-top {{ flex-direction: column; align-items: flex-start; }}
      table, thead, tbody, th, td, tr {{ display: block; }}
      thead {{ display: none; }}
      td {{
        padding: 10px 14px;
      }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1>{escape(pretty_date)}</h1>
      <p>Bankroll: EUR {bankroll:.2f}</p>
      <p>Generated for: {escape(target_date)}</p>
      <p>Source: MLB Stats API + The Odds API. JSON output for mlb.html is unchanged.</p>
    </section>
    {''.join(cards)}
    <section class="summary">
      <h2>Summary Table</h2>
      <table>
        <thead>
          <tr>
            <th>Game</th>
            <th>Pick</th>
            <th>Odds</th>
            <th>Stake %</th>
            <th>Stake EUR</th>
            <th>Return EUR</th>
            <th>Profit EUR</th>
            <th>Decision</th>
          </tr>
        </thead>
        <tbody>
          {''.join(table_rows)}
        </tbody>
      </table>
    </section>
  </div>
</body>
</html>
"""

    out_file.write_text(html, encoding="utf-8")
    return out_file


def fetch_mlb_odds(target_date: str) -> tuple[dict, str]:
    import os

    et_now = datetime.now(timezone(timedelta(hours=-4))).strftime("%Y-%m-%d %H:%M ET")
    api_key = os.environ.get("ODDS_API_KEY", "")
    if not api_key:
        print("// ODDS_API_KEY not set - skipping market odds", file=sys.stderr)
        return {}, et_now

    # Primary fetch: h2h + standard spreads — this is what the singles pipeline depends on.
    try:
        r = requests.get(
            "https://api.the-odds-api.com/v4/sports/baseball_mlb/odds",
            params={
                "apiKey": api_key,
                "regions": "uk,us",
                "markets": "h2h,spreads",
                "oddsFormat": "decimal",
                "dateFormat": "iso",
            },
            timeout=20,
        )
        remaining = r.headers.get("x-requests-remaining", "?")
        print(f"// Odds API: {r.status_code}  requests remaining: {remaining}", file=sys.stderr)
        if r.status_code != 200:
            print(f"// Odds API error: {r.text[:200]}", file=sys.stderr)
            return {}, et_now
    except Exception as e:
        print(f"// Odds API fetch failed: {e}", file=sys.stderr)
        return {}, et_now

    # Secondary fetch: alternate spreads for the acca module — optional, fails silently.
    # Kept separate so a plan/availability issue here never breaks the singles pipeline.
    alt_spreads_by_game: dict[str, list] = {}
    try:
        r_alt = requests.get(
            "https://api.the-odds-api.com/v4/sports/baseball_mlb/odds",
            params={
                "apiKey": api_key,
                "regions": "uk,us",
                "markets": "alternate_spreads",
                "oddsFormat": "decimal",
                "dateFormat": "iso",
            },
            timeout=20,
        )
        if r_alt.status_code == 200:
            for g_alt in r_alt.json():
                key = (g_alt.get("away_team", ""), g_alt.get("home_team", ""))
                alt_spreads_by_game[key] = g_alt.get("bookmakers", [])
            print(f"// Alternate spreads: fetched for {len(alt_spreads_by_game)} games", file=sys.stderr)
        else:
            print(f"// Alternate spreads not available (status {r_alt.status_code}) — acca uses standard spreads only", file=sys.stderr)
    except Exception as e:
        print(f"// Alternate spreads fetch skipped: {e}", file=sys.stderr)

    et_zone = timezone(timedelta(hours=-4))
    target = datetime.strptime(target_date, "%Y-%m-%d").replace(tzinfo=et_zone)

    result = {}
    for g in r.json():
        utc = datetime.fromisoformat(g["commence_time"].replace("Z", "+00:00"))
        et = utc.astimezone(et_zone)
        if et.date() != target.date():
            continue

        away_full = g["away_team"]
        home_full = g["home_team"]
        away_abbr = ODDS_TEAM_MAP.get(away_full)
        home_abbr = ODDS_TEAM_MAP.get(home_full)
        if not away_abbr or not home_abbr:
            continue

        def best(books, market_key, team):
            best_price = None
            best_book = None
            for bm in books:
                for mkt in bm.get("markets", []):
                    if mkt.get("key") != market_key:
                        continue
                    for outcome in mkt.get("outcomes", []):
                        if outcome.get("name") == team:
                            if best_price is None or outcome["price"] > best_price:
                                best_price = outcome["price"]
                                best_book = bm.get("key")
            return best_price, best_book

        def best_by_priority(books, market_key, team):
            preferred_books = [bm for bm in books if bm.get("key") in PREFERRED_BOOKMAKERS]
            if preferred_books:
                price, book = best(preferred_books, market_key, team)
                if price is not None:
                    return price, book

            uk_books = [bm for bm in books if bm.get("key") in UK_BOOKMAKERS]
            if uk_books:
                price, book = best(uk_books, market_key, team)
                if price is not None:
                    return price, book

            return best(books, market_key, team)

        def collect_spread_options(primary_books, alt_books, team):
            """Collect all available spread lines for a team from standard + alternate spread books."""
            best_by_line = {}
            for bm in primary_books:
                for mkt in bm.get("markets", []):
                    if mkt.get("key") != "spreads":
                        continue
                    for outcome in mkt.get("outcomes", []):
                        if outcome.get("name") != team:
                            continue
                        point = outcome.get("point")
                        price = outcome.get("price")
                        if point is None or price is None:
                            continue
                        if point not in best_by_line or price > best_by_line[point]:
                            best_by_line[point] = price
            for bm in alt_books:
                for mkt in bm.get("markets", []):
                    if mkt.get("key") != "alternate_spreads":
                        continue
                    for outcome in mkt.get("outcomes", []):
                        if outcome.get("name") != team:
                            continue
                        point = outcome.get("point")
                        price = outcome.get("price")
                        if point is None or price is None:
                            continue
                        if point not in best_by_line or price > best_by_line[point]:
                            best_by_line[point] = price
            return [{"line": k, "odds": round(v, 3)} for k, v in sorted(best_by_line.items())]

        books = g.get("bookmakers", [])
        alt_books = alt_spreads_by_game.get((away_full, home_full), [])
        away_ml, away_bk = best_moneyline(books, away_full)
        home_ml, home_bk = best_moneyline(books, home_full)
        if home_ml and away_ml:
            home_rl_point, away_rl_point = standard_run_line_points(home_ml, away_ml)
            home_rl, _ = best_spread(books, home_full, home_rl_point)
            away_rl, _ = best_spread(books, away_full, away_rl_point)
        else:
            home_rl_point = away_rl_point = None
            home_rl = away_rl = None
        away_rl_options = collect_filtered_spread_options(books, alt_books, away_full)
        home_rl_options = collect_filtered_spread_options(books, alt_books, home_full)
        home_no_vig, away_no_vig = no_vig_probs(home_ml, away_ml)

        result[(away_abbr, home_abbr)] = {
            "away_ml": away_ml,
            "home_ml": home_ml,
            "away_rl": away_rl,
            "home_rl": home_rl,
            "away_rl_point": away_rl_point,
            "home_rl_point": home_rl_point,
            "away_implied": round(1 / away_ml, 4) if away_ml else None,
            "home_implied": round(1 / home_ml, 4) if home_ml else None,
            "away_no_vig": away_no_vig,
            "home_no_vig": home_no_vig,
            "book_count": len(books),
            "away_book": away_bk,
            "home_book": home_bk,
            "away_rl_options": away_rl_options,
            "home_rl_options": home_rl_options,
        }

    print(f"// Odds API: matched {len(result)} games to today's schedule", file=sys.stderr)
    return result, et_now


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=str(date.today()))
    args = parser.parse_args()

    today = args.date
    yesterday = (datetime.strptime(today, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")

    # Read current bankroll from results log if it exists
    results_log_path = PREDICTIONS_DIR / "results_log.csv"
    if results_log_path.exists():
        import csv as _csv
        with open(results_log_path, newline="", encoding="utf-8") as _f:
            _rows = list(_csv.DictReader(_f))
        settled = [r for r in _rows if r.get("result") not in ("", "Pending")]
        if settled:
            BANKROLL_EUR = float(settled[-1]["bankroll_after"])
            print(f"// Bankroll from results log: EUR {BANKROLL_EUR:.2f}", file=sys.stderr)

    print(f"// Predictions for {today}", file=sys.stderr)
    print(f"// Loading completed games through {yesterday}...", file=sys.stderr)

    completed = fetch_completed(yesterday)
    team_state = build_team_state(completed)
    print(f"// {len(completed)} completed games, {len(team_state)} teams with history", file=sys.stderr)

    upcoming = fetch_upcoming(today)
    print(f"// {len(upcoming)} upcoming games on {today}", file=sys.stderr)

    pitchers = fetch_pitcher_stats()
    print(f"// {len(pitchers)} pitchers with {SEASON} stats", file=sys.stderr)
    bullpens = fetch_bullpen_stats(pitchers)
    bullpen_usage = fetch_recent_bullpen_usage(completed, pitchers, today)
    for team, usage in bullpen_usage.items():
        bullpens.setdefault(team, neutral_bullpen_features()).update(usage)
    print(f"// {len(bullpens)} bullpens with quality/rest features", file=sys.stderr)

    odds_map, odds_fetched_at = fetch_mlb_odds(today)

    # Archive morning odds for CLV calculation at settlement time.
    # check_movement.py will overwrite with a later snapshot if run.
    _archive_dir = PREDICTIONS_DIR / "odds_archive"
    _archive_dir.mkdir(parents=True, exist_ok=True)
    _archive_path = _archive_dir / f"{today}_morning_odds.json"
    _archive_data = {
        f"{away}_{home}": v for (away, home), v in odds_map.items()
    }
    _archive_path.write_text(
        __import__("json").dumps(_archive_data, indent=2), encoding="utf-8"
    )
    print(f"// Morning odds archived → {_archive_path.name}", file=sys.stderr)

    with open(MODEL_DIR / "moneyline_model.pkl", "rb") as f:
        saved = pickle.load(f)
    model, scaler = saved["model"], saved["scaler"]
    pkl_features = saved.get("features", FEATURES)
    print(f"// Model loaded ({len(pkl_features)} features)", file=sys.stderr)

    # Spread model — evaluated for diagnostics; production RL selection still requires validation.
    spread_model = None
    spread_model_status = "missing"
    spread_model_path = MODEL_DIR / "spread_model.pkl"
    if spread_model_path.exists():
        try:
            from mlb.scripts.spread_model import SpreadModel
            spread_model = SpreadModel.load(spread_model_path)
            spread_model_status = "loaded"
            print(
                f"// Spread model loaded (trained through {spread_model.trained_through}, "
                f"σ={spread_model.residual_std:.2f})",
                file=sys.stderr,
            )
            if spread_model.validation_passed is False:
                spread_model_status = "validation_failed"
                print(
                    f"// Spread model validation failed: {spread_model.validation_reasons}",
                    file=sys.stderr,
                )
        except Exception as _e:
            spread_model_status = "load_failed"
            print(f"// Spread model load failed (non-fatal): {_e}", file=sys.stderr)
    else:
        print(f"// Spread model missing: {spread_model_path}", file=sys.stderr)

    output = {}
    report_rows = []
    for g in upcoming:
        feat_vec, fd, has_rolling = build_features(g, team_state, pitchers, pkl_features, bullpens, bullpen_usage)
        home_prob = predict(model, scaler, feat_vec)
        away_prob = 1.0 - home_prob
        market = odds_map.get((g["away_team"], g["home_team"]))

        # If the game is already in progress, the odds returned by The Odds API are
        # live / in-play odds — driven by score, inning, and pitching changes rather
        # than pre-game market consensus.  Using them for edge calculation would produce
        # false signals.  Treat as no-market so the pick gets a SKIP / 0-edge label.
        if g.get("isLive"):
            market = None

        # Lower pick threshold from 52% to 50.5% so close games are evaluated.
        if home_prob > 0.505:
            pick_side = "home"
            model_prob = home_prob
            if market and (market.get("home_no_vig") or market.get("home_implied")):
                market_implied = market.get("home_no_vig") or market["home_implied"]
                pick_odds = market["home_ml"]
            else:
                market_implied = None
                pick_odds = None
        elif away_prob > 0.505:
            pick_side = "away"
            model_prob = away_prob
            if market and (market.get("away_no_vig") or market.get("away_implied")):
                market_implied = market.get("away_no_vig") or market["away_implied"]
                pick_odds = market["away_ml"]
            else:
                market_implied = None
                pick_odds = None
        else:
            pick_side = "none"
            model_prob = max(home_prob, away_prob)
            market_implied = None
            pick_odds = None

        edge = round(model_prob - market_implied, 4) if pick_side != "none" and market_implied is not None else 0.0

        # Close-game value flip: when the model has < 5% gap between the two teams
        # and the picked side has negative edge, check if the other side is mispriced.
        # Only applies when both probs are near 50/50 — the model has no strong conviction
        # so market pricing is the deciding factor.
        if (
            edge < 0
            and pick_side != "none"
            and market
            and abs(home_prob - away_prob) < 0.05
        ):
            other_side = "away" if pick_side == "home" else "home"
            other_prob = away_prob if other_side == "away" else home_prob
            other_impl = (market.get("away_no_vig") or market.get("away_implied")) if other_side == "away" else (market.get("home_no_vig") or market.get("home_implied"))
            other_odds = market.get("away_ml")      if other_side == "away" else market.get("home_ml")
            if other_impl and other_prob > 0.47:
                other_edge = round(other_prob - other_impl, 4)
                if other_edge > 0:
                    pick_side      = other_side
                    model_prob     = other_prob
                    market_implied = other_impl
                    pick_odds      = other_odds
                    edge           = other_edge

        # ── Spread model: compute cover probability for available lines ──────────
        # Always computed when the spread model is loaded (for inspection/logging).
        # Bet selection only switches to spread when USE_SPREAD_MODEL=True, the
        # spread model validates, and SPs are confirmed. ML inference is never
        # used for cover probability.
        spread_cover_prob = None
        spread_edge = None
        spread_point = None
        spread_odds = None
        spread_best_cover_prob = None
        spread_best_edge = None
        spread_best_point = None
        spread_best_odds = None
        spread_positive_line_count = 0
        spread_option_count = 0
        spread_rejection_reason = "not_evaluated"
        if spread_model and market and pick_side != "none":
            _sp_confirmed = (
                g.get("home_sp_name", "TBD") != "TBD"
                and g.get("away_sp_name", "TBD") != "TBD"
            )
            if not _sp_confirmed:
                spread_rejection_reason = "tbd_starting_pitcher"
            if _sp_confirmed:
                # Feature vector for the spread model (uses its own stored feature list)
                _sm_feat_list = spread_model.features or pkl_features
                _sm_vec, _, _ = build_features(g, team_state, pitchers, _sm_feat_list, bullpens, bullpen_usage)
                # Home-perspective cover prob at the standard RL point
                _rl_point = market.get("home_rl_point")
                if _rl_point is not None:
                    spread_cover_prob = round(spread_model.cover_prob(_sm_vec, float(_rl_point)), 4)
                # Find best EV spread option across all available lines
                _home_options = market.get("home_rl_options", [])
                _away_options = market.get("away_rl_options", [])
                if pick_side == "home" and _home_options:
                    _best = spread_model.best_cover_ev(
                        _sm_vec,
                        _home_options,
                        debug_label=f"{g['away_team']} @ {g['home_team']} HOME",
                        debug=SPREAD_DEBUG,
                        return_diagnostics=True,
                    )
                    if _best:
                        spread_best_edge = _best["edge"]
                        spread_best_point = _best["line"]
                        spread_best_odds = _best["odds"]
                        spread_best_cover_prob = _best["cover_prob"]
                        spread_positive_line_count = _best["positive_line_count"]
                        spread_option_count = _best["option_count"]
                        spread_rejection_reason = _best["rejection_reason"]
                        if _best["edge"] > 0:
                            spread_edge = _best["edge"]
                            spread_point = _best["line"]
                            spread_odds = _best["odds"]
                    else:
                        spread_rejection_reason = "no_valid_spread_options"
                elif pick_side == "away" and _away_options:
                    _best = spread_model.best_away_cover_ev(
                        _sm_vec,
                        _away_options,
                        debug_label=f"{g['away_team']} @ {g['home_team']} AWAY",
                        debug=SPREAD_DEBUG,
                        return_diagnostics=True,
                    )
                    if _best:
                        spread_best_edge = _best["edge"]
                        spread_best_point = _best["line"]
                        spread_best_odds = _best["odds"]
                        spread_best_cover_prob = _best["cover_prob"]
                        spread_positive_line_count = _best["positive_line_count"]
                        spread_option_count = _best["option_count"]
                        spread_rejection_reason = _best["rejection_reason"]
                        if _best["edge"] > 0:
                            spread_edge = _best["edge"]
                            spread_point = _best["line"]
                            spread_odds = _best["odds"]
                    else:
                        spread_rejection_reason = "no_valid_spread_options"
                else:
                    spread_rejection_reason = "no_spread_options"
        elif not spread_model:
            spread_rejection_reason = "spread_model_missing"
        elif not market:
            spread_rejection_reason = "market_missing"
        elif pick_side == "none":
            spread_rejection_reason = "no_ml_pick_side"

        # ── Bet selection: ML only until spread model is validated ────────────
        # Run lines are shown for context only. Do not switch to a run-line bet
        # from moneyline probability; cover probability needs its own RL model.
        rl_pick_odds = None
        use_rl = False
        if market and pick_side != "none":
            rl_pick_odds = market["home_rl"] if pick_side == "home" else market["away_rl"]

        # When USE_SPREAD_MODEL is enabled and spread model shows better EV than ML,
        # switch to the spread market. TBD starter guard is already applied above.
        spread_selection_allowed = (
            USE_SPREAD_MODEL
            and spread_model is not None
            and getattr(spread_model, "validation_passed", None) is True
        )
        if spread_selection_allowed and spread_edge is not None and spread_edge >= 0.03:
            if spread_edge > edge:
                use_rl = True
                rl_pick_odds = spread_odds
                # Re-stake using spread edge (staking ladder unchanged)
                stake = stake_tier(spread_edge, BANKROLL_EUR)
                edge = spread_edge
                spread_rejection_reason = "selected"
            else:
                spread_rejection_reason = "ml_edge_better"
        elif spread_edge is not None and spread_edge >= 0.03 and not spread_selection_allowed:
            spread_rejection_reason = "model_validation_failed"

        stake = stake_tier(edge, BANKROLL_EUR)

        h_sp = pitchers.get(g.get("home_sp_id"), {})
        a_sp = pitchers.get(g.get("away_sp_id"), {})

        row = {
            "gamePk": g["game_pk"],
            "gameStatus": "LIVE" if g.get("isLive") else "NOT_STARTED",
            "homeTeam": g["home_name"],
            "awayTeam": g["away_name"],
            "homeAbbr": g["home_team"],
            "awayAbbr": g["away_team"],
            "homeProb": round(home_prob, 4),
            "awayProb": round(away_prob, 4),
            "pickSide": pick_side,
            "modelProb": round(model_prob, 4),
            "edge": edge,
            "stake": stake,
            "hasRolling": has_rolling,
            "awayMl": market["away_ml"] if market else None,
            "homeMl": market["home_ml"] if market else None,
            "awayRl": market["away_rl"] if market else None,
            "homeRl": market["home_rl"] if market else None,
            "awayRlPoint": market.get("away_rl_point") if market else None,
            "homeRlPoint": market.get("home_rl_point") if market else None,
            "useRl": use_rl,
            "rlPickOdds": round(rl_pick_odds, 3) if use_rl and rl_pick_odds else None,
            "awayImplied": market["away_implied"] if market else None,
            "homeImplied": market["home_implied"] if market else None,
            "awayNoVig": market.get("away_no_vig") if market else None,
            "homeNoVig": market.get("home_no_vig") if market else None,
            "pickOdds": round(pick_odds, 3) if pick_odds else None,
            "marketImplied": round(market_implied, 4) if market_implied else None,
            "bookCount": market["book_count"] if market else 0,
            "hasOdds": market is not None,
            "homeBook": market["home_book"] if market else None,
            "awayBook": market["away_book"] if market else None,
            "homeSpName": g.get("home_sp_name", "TBD"),
            "awaySpName": g.get("away_sp_name", "TBD"),
            "homeSpEra": h_sp.get("era"),
            "awaySpEra": a_sp.get("era"),
            "homeSpWhip": h_sp.get("whip"),
            "awaySpWhip": a_sp.get("whip"),
            "homeSpFip": h_sp.get("fip"),
            "awaySpFip": a_sp.get("fip"),
            "homeSpBb9": h_sp.get("bb9"),
            "awaySpBb9": a_sp.get("bb9"),
            "homeSpKbbPct": h_sp.get("k_bb_pct"),
            "awaySpKbbPct": a_sp.get("k_bb_pct"),
            "homeSpHr9": h_sp.get("hr9"),
            "awaySpHr9": a_sp.get("hr9"),
            "homeSpIp": h_sp.get("ip", 0.0),
            "awaySpIp": a_sp.get("ip", 0.0),
            "homeSpW": h_sp.get("wins"),
            "homeSpL": h_sp.get("losses"),
            "awaySpW": a_sp.get("wins"),
            "awaySpL": a_sp.get("losses"),
            "homeL10WP": fd.get("HOME_L10_WIN_PCT"),
            "awayL10WP": fd.get("AWAY_L10_WIN_PCT"),
            "homeL10RD": fd.get("HOME_L10_RD"),
            "awayL10RD": fd.get("AWAY_L10_RD"),
            "eraDiff": round(fd.get("ERA_DIFF", 0), 2),
            "homeBpEra": fd.get("HOME_BP_ERA"),
            "awayBpEra": fd.get("AWAY_BP_ERA"),
            "homeBpWhip": fd.get("HOME_BP_WHIP"),
            "awayBpWhip": fd.get("AWAY_BP_WHIP"),
            "homeBpKbb": fd.get("HOME_BP_K_BB"),
            "awayBpKbb": fd.get("AWAY_BP_K_BB"),
            "homeBpIpLast3": fd.get("HOME_BP_IP_LAST_3D"),
            "awayBpIpLast3": fd.get("AWAY_BP_IP_LAST_3D"),
            "homeBpIpYesterday": fd.get("HOME_BP_IP_YESTERDAY"),
            "awayBpIpYesterday": fd.get("AWAY_BP_IP_YESTERDAY"),
            "seriesGameNumber": g.get("series_game_number"),
            "gamesInSeries": g.get("games_in_series"),
            "homeRlOptions": market.get("home_rl_options", []) if market else [],
            "awayRlOptions": market.get("away_rl_options", []) if market else [],
            # Spread model outputs (None when model not loaded or SP is TBD)
            "spreadCoverProb": spread_cover_prob,
            "spreadEdge": spread_edge,
            "spreadPoint": spread_point,
            "spreadOdds": spread_odds,
            "spreadBestCoverProb": spread_best_cover_prob,
            "spreadBestEdge": spread_best_edge,
            "spreadBestPoint": spread_best_point,
            "spreadBestOdds": spread_best_odds,
            "spreadPositiveLineCount": spread_positive_line_count,
            "spreadOptionCount": spread_option_count,
            "spreadSelectionAllowed": spread_selection_allowed,
            "spreadRejectionReason": spread_rejection_reason,
            "spreadModelLoaded": spread_model is not None,
            "spreadModelStatus": spread_model_status,
            "spreadModelValidationPassed": getattr(spread_model, "validation_passed", None) if spread_model else None,
            "spreadModelValidationReasons": getattr(spread_model, "validation_reasons", []) if spread_model else [],
            "modelFeatures": {feature: fd.get(feature) for feature in pkl_features},
        }

        output[str(g["game_pk"])] = row
        report_rows.append(row)

    accumulators      = build_accumulators(report_rows)
    md_report_path    = write_markdown_report(report_rows, today, bankroll=BANKROLL_EUR, odds_fetched_at=odds_fetched_at, accumulators=accumulators)
    excel_report_path = write_excel_report(report_rows, today, bankroll=BANKROLL_EUR)
    json_report_path  = write_json_report(report_rows, today, accumulators=accumulators)
    print(f"// Markdown report written to {md_report_path}", file=sys.stderr)
    print(f"// Excel report written to {excel_report_path}", file=sys.stderr)
    print(f"// JSON report written to {json_report_path}", file=sys.stderr)
    print(json.dumps(output, indent=2))
