"""
Build two Excel artefacts:

1. Update daily prediction XLSX files — fill Win/Loss in column F,
   re-colour every row green (Win) / red (Loss) / grey (Push) / unchanged (Skip).

2. Rebuild mlb/predictions/results_log.xlsx — cumulative daily P&L tracker
   with weekly grouping and an overall summary banner at the top.

Usage:
    python mlb/scripts/build_tracker_xlsx.py
    python mlb/scripts/build_tracker_xlsx.py --date 2026-04-14
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ── paths ──────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))
from mlb.scripts.predict_today import format_prediction_date, PREDICTIONS_DIR, ODDS_TEAM_MAP
from mlb.scripts.record_results import read_log

RESULTS_CSV         = PREDICTIONS_DIR / "results_log.csv"
RESULTS_CSV_UPDATED = PREDICTIONS_DIR / "results_log_updated.csv"
RESULTS_XLSX        = PREDICTIONS_DIR / "results_log.xlsx"

# ── colour palette ─────────────────────────────────────────────────────────────
def fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)

WIN_FILL       = fill("C6EFCE")   # soft green
LOSS_FILL      = fill("FFC7CE")   # soft red
PUSH_FILL      = fill("D9E1F2")   # soft blue
SKIP_FILL      = fill("FFF1EC")   # pale peach (existing default)
BET_FILL       = fill("EAF8F1")   # pale teal  (existing default)

BANNER_FILL    = fill("183153")   # deep navy
HEADER_FILL    = fill("244063")   # mid navy
WEEK_FILL      = fill("E2EFDA")   # pale green for weekly totals
SUB_FILL       = fill("E7F0FF")   # pale blue for sub-rows / labels
SUMMARY_FILL   = fill("1F4E79")   # darker navy for overall summary

WHITE_FONT     = Font(color="FFFFFF", bold=True)
BOLD_FONT      = Font(bold=True)
PLAIN_FONT     = Font()

def thin_border(hex_colour: str = "D0D7E2") -> Border:
    s = Side(style="thin", color=hex_colour)
    return Border(left=s, right=s, top=s, bottom=s)

BORDER = thin_border()


# ── helpers ───────────────────────────────────────────────────────────────────

def load_results() -> list[dict]:
    return read_log(RESULTS_CSV)


def _norm(value) -> str:
    return str(value or "").strip().lower()


def _team_texts(abbr: str) -> set[str]:
    abbr_text = str(abbr or "").strip()
    full = ODDS_TEAM_MAP.get(abbr_text, "")
    return {t.lower() for t in (abbr_text, full) if t}


def _result_indexes(day_rows: list[dict]) -> tuple[dict[str, dict], dict[tuple[str, str, str], dict]]:
    by_pk = {}
    by_team_pick = {}
    for row in day_rows:
        game_pk = str(row.get("game_pk") or "").strip()
        if game_pk:
            by_pk[game_pk] = row
        home = str(row.get("home_team") or "").strip()
        away = str(row.get("away_team") or "").strip()
        pick = str(row.get("pick_team") or "").strip()
        if home and away and pick:
            by_team_pick[tuple(sorted((home, away))) + (pick,)] = row
    return by_pk, by_team_pick


def _find_result_for_xlsx_row(ws, row_idx: int, by_pk: dict[str, dict], by_team_pick: dict[tuple[str, str, str], dict]) -> dict | None:
    game_pk = str(ws.cell(row=row_idx, column=8).value or "").strip()
    if game_pk and game_pk in by_pk:
        return by_pk[game_pk]

    game_text = _norm(ws.cell(row=row_idx, column=1).value)
    pick_text = _norm(ws.cell(row=row_idx, column=2).value)
    for (team_a, team_b, pick_team), result in by_team_pick.items():
        team_a_match = any(text in game_text for text in _team_texts(team_a))
        team_b_match = any(text in game_text for text in _team_texts(team_b))
        pick_match = any(text in pick_text for text in _team_texts(pick_team))
        if team_a_match and team_b_match and pick_match:
            return result
    return None


def iso_week_label(date_str: str) -> str:
    """Return the ISO week label: 'Week N (Mon DD – Sun DD Mmm YYYY)'"""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    monday = dt - timedelta(days=dt.weekday())
    sunday = monday + timedelta(days=6)
    return f"Week {dt.isocalendar()[1]}  ({monday.strftime('%-d %b')} – {sunday.strftime('%-d %b %Y')})"


def daily_stats(rows: list[dict]) -> dict:
    bets   = [r for r in rows if r["decision"] == "BET"]
    wins   = [r for r in bets if r["result"] == "Win"]
    losses = [r for r in bets if r["result"] == "Loss"]
    pushes = [r for r in bets if r["result"] == "Push"]
    staked = sum(float(r["stake_eur"]) for r in bets)
    pnl    = sum(float(r["pnl"]) for r in bets)
    # bankroll at end of day = bankroll_after of last settled bet (or before of first row)
    settled = [r for r in bets if r["result"] not in ("", "N/A", "Pending")]
    if settled:
        bankroll_end = float(settled[-1]["bankroll_after"])
    else:
        non_empty = [r for r in rows if r["bankroll_before"] not in ("", None)]
        bankroll_end = float(non_empty[0]["bankroll_before"]) if non_empty else 0.0
    return {
        "bets": len(bets),
        "wins": len(wins),
        "losses": len(losses),
        "pushes": len(pushes),
        "staked": staked,
        "pnl": pnl,
        "bankroll_end": bankroll_end,
    }


# ── 1. Update prediction XLSX for a given date ─────────────────────────────────

def update_prediction_xlsx(target_date: str, all_results: list[dict]) -> None:
    month_folder, day_folder, file_stub = format_prediction_date(target_date)
    xlsx_path = PREDICTIONS_DIR / month_folder / day_folder / f"{file_stub}.xlsx"
    if not xlsx_path.exists():
        print(f"// Prediction xlsx not found: {xlsx_path}", file=sys.stderr)
        return

    day_rows = [r for r in all_results if r["date"] == target_date]
    if not day_rows:
        print(f"// No results in CSV for {target_date}", file=sys.stderr)
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active

    HEADER_ROW = 6
    data_start = HEADER_ROW + 1

    # Collect data rows (rows that have content in column A)
    data_row_indices = []
    for row_idx in range(data_start, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is None:
            break
        data_row_indices.append(row_idx)

    by_pk, by_team_pick = _result_indexes(day_rows)
    for xlsx_row_idx in data_row_indices:
        r = _find_result_for_xlsx_row(ws, xlsx_row_idx, by_pk, by_team_pick)
        if r is None:
            continue
        decision = r.get("decision", "SKIP")
        result   = r.get("result", "N/A")

        if decision == "BET":
            # Write result into column F
            ws.cell(row=xlsx_row_idx, column=6).value = result

            if result == "Win":
                row_fill = WIN_FILL
            elif result == "Loss":
                row_fill = LOSS_FILL
            elif result == "Push":
                row_fill = PUSH_FILL
            else:
                row_fill = BET_FILL   # pending / unknown
        else:
            row_fill = SKIP_FILL

        # Re-colour every cell A–G on this row
        for col in range(1, 8):
            ws.cell(row=xlsx_row_idx, column=col).fill = row_fill

    wb.save(xlsx_path)
    print(f"// Updated prediction xlsx: {xlsx_path}", file=sys.stderr)


def update_updated_prediction_xlsx(target_date: str) -> None:
    """Same as update_prediction_xlsx but targets the (Updated).xlsx using results_log_updated.csv."""
    if not RESULTS_CSV_UPDATED.exists():
        return

    month_folder, day_folder, file_stub = format_prediction_date(target_date)
    xlsx_path = PREDICTIONS_DIR / month_folder / day_folder / f"{file_stub} (Updated).xlsx"
    if not xlsx_path.exists():
        return

    all_results = read_log(RESULTS_CSV_UPDATED)

    day_rows = [r for r in all_results if r["date"] == target_date]
    if not day_rows:
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active

    HEADER_ROW = 6
    data_start = HEADER_ROW + 1

    data_row_indices = []
    for row_idx in range(data_start, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is None:
            break
        data_row_indices.append(row_idx)

    by_pk, by_team_pick = _result_indexes(day_rows)
    for xlsx_row_idx in data_row_indices:
        r = _find_result_for_xlsx_row(ws, xlsx_row_idx, by_pk, by_team_pick)
        if r is None:
            continue
        decision = r.get("decision", "SKIP")
        result   = r.get("result", "N/A")

        if decision == "BET":
            ws.cell(row=xlsx_row_idx, column=6).value = result
            if result == "Win":
                row_fill = WIN_FILL
            elif result == "Loss":
                row_fill = LOSS_FILL
            elif result == "Push":
                row_fill = PUSH_FILL
            else:
                row_fill = BET_FILL
        else:
            row_fill = SKIP_FILL

        for col in range(1, 8):
            ws.cell(row=xlsx_row_idx, column=col).fill = row_fill

    wb.save(xlsx_path)
    print(f"// Updated (Updated) prediction xlsx: {xlsx_path}", file=sys.stderr)


# ── 2. Build the cumulative results_log.xlsx ───────────────────────────────────

DETAIL_SUBHDR_FILL = fill("2C5F8A")  # slightly lighter navy for the detail sub-header
DETAIL_SKIP_FILL   = fill("F0F0F0")  # neutral grey for skips in detail

# Detail column headers (shown when a day row is expanded)
DETAIL_HEADERS = ["  Game", "Pick", "Odds", "Stake (EUR)", "Decision", "Result", "P&L (EUR)", "Bankroll (EUR)"]


def _write_cell(ws, row_idx: int, col: int, value, *, f=None, bg=None,
                num_fmt: str = None, align: Alignment = None, border: Border = None):
    c = ws.cell(row=row_idx, column=col, value=value)
    if f:       c.font        = f
    if bg:      c.fill        = bg
    if num_fmt: c.number_format = num_fmt
    if align:   c.alignment   = align
    if border:  c.border      = border
    return c


def build_results_xlsx(all_results: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Tracker"

    # Summary rows sit ABOVE their detail rows — Excel must know this
    ws.sheet_properties.outlinePr.summaryBelow = False

    # ── column widths ─────────────────────────────────────────────────────────
    # A: Date / Game          B: Bets / Pick
    # C: Won / Odds           D: Lost / Stake
    # E: Win% / Decision      F: Staked / Result
    # G: P&L                  H: Bankroll
    COL_WIDTHS = {"A": 36, "B": 18, "C": 10, "D": 13, "E": 11, "F": 11, "G": 14, "H": 16}
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    def set_row(row_idx, height, outline_level=0, hidden=False):
        rd = ws.row_dimensions[row_idx]
        rd.height        = height
        rd.outline_level = outline_level
        rd.hidden        = hidden

    CENTER   = Alignment(horizontal="center", vertical="center")
    LEFT_IN2 = Alignment(horizontal="left",   vertical="center", indent=2)
    LEFT_IN4 = Alignment(horizontal="left",   vertical="center", indent=4)
    EUR_FMT  = '"EUR" #,##0.00'
    EUR_FMT2 = '"EUR" #,##0.00;[Red]"-EUR" #,##0.00'  # negative in red

    # ── overall stats ─────────────────────────────────────────────────────────
    all_bets    = [r for r in all_results if r["decision"] == "BET"]
    all_wins    = [r for r in all_bets if r["result"] == "Win"]
    all_losses  = [r for r in all_bets if r["result"] == "Loss"]
    all_pnl     = sum(float(r["pnl"])       for r in all_bets)
    all_staked  = sum(float(r["stake_eur"]) for r in all_bets)
    starting_br = 500.0
    current_br  = starting_br + all_pnl
    win_rate    = (len(all_wins) / len(all_bets) * 100) if all_bets else 0.0
    roi         = (all_pnl / all_staked * 100) if all_staked else 0.0

    # ── Banner row ────────────────────────────────────────────────────────────
    row = 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "MLB Betting P&L Tracker"
    ws[f"A{row}"].font      = Font(color="FFFFFF", bold=True, size=15)
    ws[f"A{row}"].fill      = BANNER_FILL
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    set_row(row, 24)

    # ── Starting bankroll row ─────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = f"Starting Bankroll: EUR {starting_br:,.2f}"
    ws[f"A{row}"].font      = BOLD_FONT
    ws[f"A{row}"].fill      = SUB_FILL
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    set_row(row, 16)

    # ── Live summary stat row ─────────────────────────────────────────────────
    row += 1
    pnl_sign = "+" if all_pnl >= 0 else ""
    stat_line = (
        f"Bankroll: EUR {current_br:,.2f}"
        f"    |    P&L: {pnl_sign}EUR {all_pnl:,.2f}"
        f"    |    ROI: {roi:+.1f}%"
        f"    |    Win Rate: {win_rate:.1f}%  ({len(all_wins)}W / {len(all_losses)}L)"
        f"    |    Total Bets: {len(all_bets)}"
    )
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = stat_line
    ws[f"A{row}"].font      = Font(color="FFFFFF", bold=True, size=10)
    ws[f"A{row}"].fill      = SUMMARY_FILL
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    set_row(row, 18)

    # ── Thin spacer ───────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"].fill = SUB_FILL
    set_row(row, 5)

    # ── Column headers (day-level) ────────────────────────────────────────────
    row += 1
    DAY_HEADERS = ["Date", "Bets", "Won", "Lost", "Win %", "Staked (EUR)", "P&L (EUR)", "Bankroll (EUR)"]
    for col, hdr in enumerate(DAY_HEADERS, start=1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font      = Font(color="FFFFFF", bold=True)
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = BORDER
    set_row(row, 18)
    HEADER_ROW_IDX = row

    # ── Group rows by date, then by ISO week ──────────────────────────────────
    by_date: dict[str, list[dict]] = defaultdict(list)
    for r in all_results:
        by_date[r["date"]].append(r)

    sorted_dates = sorted(by_date.keys())

    def iso_week_key(date_str: str) -> str:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        y, w, _ = dt.isocalendar()
        return f"{y}-W{w:02d}"

    by_week: dict[str, list[str]] = defaultdict(list)
    for d in sorted_dates:
        by_week[iso_week_key(d)].append(d)

    # ── Write weeks ───────────────────────────────────────────────────────────
    for week_key, week_dates in sorted(by_week.items()):
        first_dt = datetime.strptime(week_dates[0], "%Y-%m-%d")
        monday   = first_dt - timedelta(days=first_dt.weekday())
        sunday   = monday + timedelta(days=6)
        week_num = first_dt.isocalendar()[1]
        week_label = (
            f"Week {week_num}  "
            f"({monday.strftime('%d %b')} – {sunday.strftime('%d %b %Y')})"
        )

        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = week_label
        ws[f"A{row}"].font      = Font(bold=True, size=10, color="183153")
        ws[f"A{row}"].fill      = fill("B8D4E8")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws[f"A{row}"].border    = BORDER
        set_row(row, 16)

        week_stats = {"bets": 0, "wins": 0, "losses": 0, "staked": 0.0, "pnl": 0.0}

        for date_str in week_dates:
            day_rows = by_date[date_str]
            s        = daily_stats(day_rows)
            week_stats["bets"]   += s["bets"]
            week_stats["wins"]   += s["wins"]
            week_stats["losses"] += s["losses"]
            week_stats["staked"] += s["staked"]
            week_stats["pnl"]    += s["pnl"]

            win_pct = (s["wins"] / s["bets"] * 100) if s["bets"] else 0.0
            dt      = datetime.strptime(date_str, "%Y-%m-%d")

            day_summary_fill = (
                WIN_FILL  if s["pnl"] > 0 else
                LOSS_FILL if s["pnl"] < 0 else SUB_FILL
            )

            # ── Day summary row (outline_level 0, always visible) ─────────────
            row += 1
            day_summary_row = row
            day_values = [
                dt.strftime("%a %d %b %Y"),
                s["bets"], s["wins"], s["losses"],
                f"{win_pct:.0f}%",
                s["staked"], s["pnl"], s["bankroll_end"],
            ]
            for col, val in enumerate(day_values, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill      = day_summary_fill
                c.border    = BORDER
                c.alignment = CENTER
                if col in (6, 7, 8): c.number_format = EUR_FMT
                if col == 7:         c.font = Font(bold=True)
            ws.cell(row=row, column=1).alignment = LEFT_IN2
            set_row(row, 16)

            # ── Detail sub-header row (level 1, visible by default) ─────────
            row += 1
            for col, hdr in enumerate(DETAIL_HEADERS, start=1):
                c = ws.cell(row=row, column=col, value=hdr)
                c.font      = Font(color="FFFFFF", bold=True, size=9)
                c.fill      = DETAIL_SUBHDR_FILL
                c.border    = BORDER
                c.alignment = CENTER if col > 1 else Alignment(horizontal="left", vertical="center", indent=4)
            set_row(row, 13, outline_level=1, hidden=False)

            # ── Individual bet / skip rows (level 1, hidden/collapsed) ───────
            for r in day_rows:
                row += 1
                decision = r.get("decision", "SKIP")
                result   = r.get("result", "N/A")

                if decision == "BET":
                    if result == "Win":   det_fill = WIN_FILL
                    elif result == "Loss": det_fill = LOSS_FILL
                    elif result == "Push": det_fill = PUSH_FILL
                    else:                  det_fill = BET_FILL
                else:
                    det_fill = DETAIL_SKIP_FILL

                try:   odds_val = float(r["pick_odds"])
                except (ValueError, TypeError): odds_val = None
                try:   stake_val = float(r["stake_eur"])
                except (ValueError, TypeError): stake_val = 0.0
                try:   pnl_val = float(r["pnl"])
                except (ValueError, TypeError): pnl_val = 0.0
                try:   br_val = float(r["bankroll_after"])
                except (ValueError, TypeError): br_val = None

                game_label = f"  {r['away_team']} @ {r['home_team']}"
                det_values = [
                    game_label,
                    r.get("pick_team", ""),
                    odds_val,
                    stake_val,
                    decision,
                    result if result != "N/A" else "—",
                    pnl_val if decision == "BET" else None,
                    br_val  if decision == "BET" else None,
                ]
                for col, val in enumerate(det_values, start=1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.fill      = det_fill
                    c.border    = BORDER
                    c.font      = Font(size=9)
                    c.alignment = CENTER
                    if col in (4, 7, 8) and val is not None:
                        c.number_format = EUR_FMT
                    if col == 3 and val is not None:
                        c.number_format = "0.00"
                    if col == 6:
                        if result == "Win":
                            c.font = Font(size=9, bold=True, color="276221")
                        elif result == "Loss":
                            c.font = Font(size=9, bold=True, color="9C0006")
                ws.cell(row=row, column=1).alignment = LEFT_IN4
                set_row(row, 13, outline_level=1, hidden=False)

        # ── Week totals row ───────────────────────────────────────────────────
        row += 1
        week_win_pct = (week_stats["wins"] / week_stats["bets"] * 100) if week_stats["bets"] else 0.0
        week_pnl     = week_stats["pnl"]
        wk_fill = fill("AADDAA") if week_pnl > 0 else (fill("FFAAAA") if week_pnl < 0 else WEEK_FILL)

        week_vals = [
            f"Week {week_num} Total",
            week_stats["bets"], week_stats["wins"], week_stats["losses"],
            f"{week_win_pct:.0f}%",
            week_stats["staked"], week_pnl, "",
        ]
        for col, val in enumerate(week_vals, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = BOLD_FONT
            c.fill      = wk_fill
            c.border    = BORDER
            c.alignment = CENTER
            if col in (6, 7): c.number_format = EUR_FMT
        ws.cell(row=row, column=1).alignment = LEFT_IN2
        set_row(row, 15)

        # thin spacer between weeks
        row += 1
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = fill("F0F0F0")
        set_row(row, 4)

    ws.freeze_panes = f"A{HEADER_ROW_IDX + 1}"

    wb.save(RESULTS_XLSX)
    print(f"// Results log xlsx written: {RESULTS_XLSX}", file=sys.stderr)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build prediction + P&L tracker XLSXs")
    parser.add_argument("--date", help="Update prediction xlsx for this date only (YYYY-MM-DD)")
    args = parser.parse_args()

    all_results = load_results()
    if not all_results:
        print("No results in results_log.csv yet.", file=sys.stderr)
        sys.exit(0)

    # 1. Update prediction xlsx files
    if args.date:
        dates_to_update = [args.date]
    else:
        dates_to_update = sorted(set(r["date"] for r in all_results))

    for d in dates_to_update:
        update_prediction_xlsx(d, all_results)
        update_updated_prediction_xlsx(d)

    # 2. Rebuild results_log.xlsx
    build_results_xlsx(all_results)
    print("Done.", file=sys.stderr)


if __name__ == "__main__":
    main()
